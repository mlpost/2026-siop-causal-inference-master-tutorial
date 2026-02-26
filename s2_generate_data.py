
"""
Mock Data Generator for Leadership Development Program - Scenario 2
====================================================================

This script generates a deterministic mock dataset for a statistics tutorial
on causal inference using an observational (self-selection) design.

Design: Open enrollment - voluntary participation with self-selection bias
- Treated:  ~500  managers who voluntarily enrolled (Jan-Mar training)
- Control: ~8,500 managers who did not participate
- Self-selection driven by Organization and performance rating
- No Below / Far Below performers are treated
- ~25 % of managers are new (no prior manager-level baseline scores)

Key differences vs Scenario 1 (staggered RCT):
- Treatment is NOT randomised - covariates are imbalanced
- Baseline (prior-year) engagement scores are included and act as confounders
- Workshop teaches IPTW / Prognostic Score adjustment rather than simple GEE

Author: Generated for Statistics Tutorial
Date: 2026
"""

import numpy as np
import pandas as pd
from scipy import stats
from scipy.special import expit  # logistic function
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# SECTION 1: SETUP AND CONSTANTS
# ============================================================================

SEED = 42
np.random.seed(SEED)

# Sample sizes
N_TOTAL = 9000
N_TREATED_TARGET = 500          # approximate - calibrated via propensity model
MIN_TEAM_SIZE = 5
MAX_TEAM_SIZE = 12

# Demographics categories (same as S1)
REGIONS = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East & Africa']
ORGANIZATIONS = ['R&D', 'Commercial', 'Manufacturing', 'Digital', 'HR', 'Finance']
JOB_FAMILIES = [
    'Clinical Operations', 'Regulatory Affairs', 'Data Science', 'Medical Affairs',
    'Sales', 'Marketing', 'Supply Chain', 'Quality Assurance', 'Pharmacovigilance',
    'Market Access', 'Human Resources', 'IT & Digital', 'Finance & Accounting',
    'Legal & Compliance', 'Communications'
]
PERFORMANCE_RATINGS = ['Far Below', 'Below', 'Meets', 'Exceeds', 'Far Exceeds']
GENDERS = ['Male', 'Female', 'Non-Binary/Other']

# Survey scale parameters (1-5 Likert)
SURVEY_MIN = 1
SURVEY_MAX = 5

print("=" * 80)
print("LEADERSHIP DEVELOPMENT PROGRAM - SCENARIO 2 MOCK DATA GENERATOR")
print("=" * 80)
print(f"\nSeed: {SEED}")
print(f"Total Managers: {N_TOTAL} (target ~{N_TREATED_TARGET} treated)")
print(f"Expected Direct Reports: ~{N_TOTAL * 8} (team sizes {MIN_TEAM_SIZE}-{MAX_TEAM_SIZE})")

# ============================================================================
# SECTION 2: GENERATE MANAGER-LEVEL DEMOGRAPHICS
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING MANAGER-LEVEL DATA")
print("=" * 80)

manager_ids = np.arange(1, N_TOTAL + 1)

# Demographics - same distributional assumptions as S1 but for the full population
regions = np.random.choice(REGIONS, size=N_TOTAL)
organizations = np.random.choice(ORGANIZATIONS, size=N_TOTAL)
job_families = np.random.choice(JOB_FAMILIES, size=N_TOTAL)

perf_probs = [0.05, 0.15, 0.60, 0.15, 0.05]
performance = np.random.choice(PERFORMANCE_RATINGS, size=N_TOTAL, p=perf_probs)

gender_probs = [0.48, 0.48, 0.04]
genders = np.random.choice(GENDERS, size=N_TOTAL, p=gender_probs)

ages = np.random.normal(38, 6, N_TOTAL).clip(28, 55).round().astype(int)
tenures = np.random.gamma(3, 4, N_TOTAL).clip(1, 36).round().astype(int)

print(f"[OK] Generated demographics for {N_TOTAL} managers")

# ============================================================================
# SECTION 3: NEW MANAGER FLAG
# ============================================================================

print("\nGenerating new-manager flag (~25 %)...")

# New managers are more likely among short-tenure individuals.
# P(new_manager) = logistic( intercept + slope * (12 - tenure) )   ~25 % overall
prob_new = expit(-1.3 + 0.12 * (12 - tenures))
is_new_manager = (np.random.uniform(size=N_TOTAL) < prob_new).astype(int)

pct_new = is_new_manager.mean() * 100
print(f"[OK] {is_new_manager.sum()} managers flagged as new ({pct_new:.1f} %)")

# ============================================================================
# SECTION 4: GENERATE PRIOR-YEAR BASELINE SCORES
# ============================================================================

print("\nGenerating prior-year baseline scores...")

def generate_baseline(base_mean, base_sd, n):
    """Generate a baseline score (1-5) as an independent draw."""
    scores = np.random.normal(base_mean, base_sd, n)
    return np.clip(np.round(scores, 1), SURVEY_MIN, SURVEY_MAX)

# Manager-level baselines (NaN for new managers)
baseline_manager_efficacy_raw = generate_baseline(3.3, 0.85, N_TOTAL)

baseline_manager_efficacy = np.where(is_new_manager == 1, np.nan, baseline_manager_efficacy_raw)

# Individual-level baselines (all managers have these - measured as IC before promotion)
baseline_workload           = generate_baseline(3.1, 0.95, N_TOTAL)
baseline_turnover_intention = generate_baseline(2.7, 1.00, N_TOTAL)

print(f"  baseline_manager_efficacy : {np.nanmean(baseline_manager_efficacy):.2f} (NaN count: {np.isnan(baseline_manager_efficacy).sum()})")
print(f"  baseline_workload         : {baseline_workload.mean():.2f}")
print(f"  baseline_turnover_intention: {baseline_turnover_intention.mean():.2f}")

# ============================================================================
# SECTION 5: SELF-SELECTION INTO TREATMENT
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING SELF-SELECTION (TREATMENT ASSIGNMENT)")
print("=" * 80)

# Selection drivers: Organization + Performance Rating
# R&D and Digital promoted the programme more heavily - higher participation
org_weight = np.zeros(N_TOTAL)
org_weight[organizations == 'R&D']           =  0.90
org_weight[organizations == 'Digital']       =  0.75
org_weight[organizations == 'Commercial']    =  0.0
org_weight[organizations == 'Manufacturing'] = -0.20
org_weight[organizations == 'HR']            = -0.10
org_weight[organizations == 'Finance']       = -0.25

# Performance rating effect - higher performers are more likely to self-select
# Below and Far Below are hard-blocked from treatment (no treated managers)
perf_weight = np.zeros(N_TOTAL)
perf_weight[performance == 'Far Exceeds'] =  1.20
perf_weight[performance == 'Exceeds']     =  0.60
perf_weight[performance == 'Meets']       =  0.0
perf_weight[performance == 'Below']       = -20.0   # effectively blocks treatment
perf_weight[performance == 'Far Below']   = -20.0   # effectively blocks treatment

# Calibrate intercept so that ~N_TREATED_TARGET managers are treated
# Use bisection for precise calibration
lo, hi = -6.0, 2.0
for _ in range(200):
    intercept = (lo + hi) / 2
    logit_p = intercept + org_weight + perf_weight
    p_treat = expit(logit_p)
    expected_treated = p_treat.sum()
    if expected_treated < N_TREATED_TARGET:
        lo = intercept
    else:
        hi = intercept
    if abs(expected_treated - N_TREATED_TARGET) < 0.5:
        break

print(f"  Calibrated intercept: {intercept:.4f}")
print(f"  Expected treated (sum of probabilities): {p_treat.sum():.1f}")

# Draw treatment from Bernoulli(p_treat)
treatment = (np.random.uniform(size=N_TOTAL) < p_treat).astype(int)

# Hard-enforce: no Below or Far Below treated
treatment[(performance == 'Below') | (performance == 'Far Below')] = 0

n_treated = treatment.sum()
n_control = N_TOTAL - n_treated
print(f"[OK] Treatment assigned: {n_treated} treated, {n_control} control")

# Store propensity scores for diagnostics
propensity_scores = p_treat

# ============================================================================
# SECTION 6: GENERATE MANAGER OUTCOMES
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING MANAGER OUTCOMES")
print("=" * 80)

def generate_outcome_with_baseline(base_mean, base_sd, treatment_effect_d,
                                   treatment, baseline, baseline_r=0.50,
                                   hetero_mask=None, hetero_extra_d=0.0):
    """
    Generate an outcome (1-5) that is correlated with its baseline score
    AND has an independent treatment effect on top.

    Parameters
    ----------
    base_mean       : control-group mean (for units at baseline mean)
    base_sd         : residual SD of outcome
    treatment_effect_d : Cohen's d for the treatment effect
    treatment       : 0/1 treatment indicator
    baseline        : prior-year score (may contain NaN)
    baseline_r      : year-over-year correlation (outcome | baseline)
    hetero_mask     : boolean mask for subgroup with extra effect
    hetero_extra_d  : additional Cohen's d for the subgroup
    """
    n = len(treatment)

    # Replace NaN baselines with the population mean of non-NaN values
    bl_filled = np.where(np.isnan(baseline), np.nanmean(baseline), baseline)
    bl_centered = bl_filled - np.nanmean(baseline)

    # Outcome = baseline component + treatment effect + noise
    baseline_component = baseline_r * bl_centered
    treatment_effect = treatment_effect_d * base_sd * treatment

    # Heterogeneous treatment effect (e.g. stronger for R&D)
    if hetero_mask is not None:
        treatment_effect = treatment_effect + hetero_extra_d * base_sd * treatment * hetero_mask

    residual_sd = base_sd * np.sqrt(1 - baseline_r ** 2)
    noise = np.random.normal(0, residual_sd, n)

    scores = base_mean + baseline_component + treatment_effect + noise
    scores = np.clip(np.round(scores, 1), SURVEY_MIN, SURVEY_MAX)
    return scores


# R&D heterogeneity mask
is_rnd = (organizations == 'R&D').astype(float)

# --- Manager Efficacy Index: d  0.50, significant, moderate  (R&D: +0.15)
manager_efficacy = generate_outcome_with_baseline(
    base_mean=3.4, base_sd=0.90, treatment_effect_d=0.50,
    treatment=treatment, baseline=baseline_manager_efficacy_raw,
    baseline_r=0.50,
    hetero_mask=is_rnd, hetero_extra_d=0.15
)
print("  [OK] manager_efficacy_index  (d0.50, R&D extra +0.15)")

# --- Workload Index (Manager): d  0.02, non-significant
# (Training doesn't meaningfully affect perceived workload)
workload_manager = generate_outcome_with_baseline(
    base_mean=3.2, base_sd=1.00, treatment_effect_d=0.02,
    treatment=treatment, baseline=baseline_workload,
    baseline_r=0.45
)
print("  [OK] workload_index_mgr      (d0.02, ns)")

# --- Turnover Intention Index (Manager): d  0.25, significant, small
#     New managers get extra +0.20 d (total ~0.45 for new managers)
turnover_intention_manager = generate_outcome_with_baseline(
    base_mean=2.8, base_sd=1.00, treatment_effect_d=0.25,
    treatment=treatment, baseline=baseline_turnover_intention,
    baseline_r=0.50,
    hetero_mask=is_new_manager.astype(float), hetero_extra_d=0.20
)
print("  [OK] turnover_intention_mgr  (d0.25, new mgr extra +0.20)")

# ============================================================================
# SECTION 7: GENERATE MANAGER RETENTION OUTCOMES
# ============================================================================

print("\nGenerating manager retention outcomes...")

def generate_retention_with_baseline(base_rate, treatment_or, treatment,
                                     baseline_turnover_intent,
                                     hetero_mask=None, hetero_extra_or=1.0):
    """
    Generate binary retention with treatment OR and a baseline covariate effect.
    Higher baseline turnover_intention (= more intention to stay) -> higher retention.

    Parameters
    ----------
    hetero_mask     : boolean/int mask for subgroup with extra effect
    hetero_extra_or : multiplicative OR boost for the subgroup
                      e.g. 1.5 means the subgroup OR = treatment_or * 1.5
    """
    base_logit = np.log(base_rate / (1 - base_rate))
    treat_logit = np.log(treatment_or)

    # Heterogeneous treatment effect for subgroup (additive on log-odds scale)
    hetero_logit = 0.0
    if hetero_mask is not None:
        hetero_logit = np.log(hetero_extra_or) * hetero_mask

    # Centre baseline and add as linear predictor (odds-scale influence)
    bl_centered = baseline_turnover_intent - baseline_turnover_intent.mean()
    bl_effect = 0.30 * bl_centered  # moderate baseline influence

    logits = base_logit + (treat_logit + hetero_logit) * treatment + bl_effect
    probs = expit(logits)
    return (np.random.uniform(size=len(treatment)) < probs).astype(int)

# New-manager heterogeneity mask for retention/turnover
is_new = is_new_manager.astype(float)

# 3-month: OR  2.0 (significant, moderate); new managers OR - 1.5  3.0
retention_3mo = generate_retention_with_baseline(
    base_rate=0.90, treatment_or=2.0, treatment=treatment,
    baseline_turnover_intent=baseline_turnover_intention,
    hetero_mask=is_new, hetero_extra_or=1.5
)

# 6-month: CONDITIONAL on surviving to 3 months
# target cumulative: treated ~93 %, control ~86 %
p_surv_6_treated = 0.93 / 0.95
p_surv_6_control = 0.86 / 0.90
cond_prob_6 = np.where(treatment == 1, p_surv_6_treated, p_surv_6_control)
retention_6mo = retention_3mo * (np.random.binomial(1, cond_prob_6, N_TOTAL))

# 9-month: conditional
# target cumulative: treated ~91 %, control ~83 %
p_surv_9_treated = 0.91 / 0.93
p_surv_9_control = 0.83 / 0.86
cond_prob_9 = np.where(treatment == 1, p_surv_9_treated, p_surv_9_control)
retention_9mo = retention_6mo * (np.random.binomial(1, cond_prob_9, N_TOTAL))

# 12-month: conditional
# target cumulative: treated ~89 %, control ~80 %
p_surv_12_treated = 0.89 / 0.91
p_surv_12_control = 0.80 / 0.83
cond_prob_12 = np.where(treatment == 1, p_surv_12_treated, p_surv_12_control)
retention_12mo = retention_9mo * (np.random.binomial(1, cond_prob_12, N_TOTAL))

# Monotonicity checks
assert (retention_6mo  <= retention_3mo).all(),  "ERROR: non-monotonic 36 mo!"
assert (retention_9mo  <= retention_6mo).all(),  "ERROR: non-monotonic 69 mo!"
assert (retention_12mo <= retention_9mo).all(),  "ERROR: non-monotonic 912 mo!"
print("  [OK] Retention monotonicity verified")

# ============================================================================
# SECTION 8: ASSEMBLE MANAGER DATAFRAME
# ============================================================================

# Generate team sizes here so they can be included in manager df AND used for DR expansion
team_sizes = np.random.randint(MIN_TEAM_SIZE, MAX_TEAM_SIZE + 1, N_TOTAL)

df_managers = pd.DataFrame({
    'manager_id': manager_ids,
    'treatment': treatment,
    'region': regions,
    'organization': organizations,
    'job_family': job_families,
    'performance_rating': performance,
    'gender': genders,
    'age': ages,
    'tenure_months': tenures,
    'is_new_manager': is_new_manager,
    'team_size': team_sizes,
    'baseline_manager_efficacy': baseline_manager_efficacy,
    'baseline_workload': baseline_workload,
    'baseline_turnover_intention': baseline_turnover_intention,
    'propensity_score': np.round(propensity_scores, 4),
    'retention_3month': retention_3mo,
    'retention_6month': retention_6mo,
    'retention_9month': retention_9mo,
    'retention_12month': retention_12mo,
    'manager_efficacy_index': manager_efficacy,
    'workload_index_mgr': workload_manager,
    'turnover_intention_index_mgr': turnover_intention_manager,
})

print(f"\n[OK] Manager dataframe assembled: {df_managers.shape}")

# ============================================================================
# SECTION 9: GENERATE DIRECT REPORT DATA
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING DIRECT REPORT DATA")
print("=" * 80)

# team_sizes already generated in Section 8 (same array used here for DR expansion)
total_dr = team_sizes.sum()

print(f"Total direct reports: {total_dr}")
print(f"Average team size: {team_sizes.mean():.1f}")
print(f"  Treated DRs : ~{team_sizes[treatment == 1].sum()}")
print(f"  Control DRs : ~{team_sizes[treatment == 0].sum()}")

# Repeat manager-level info for each DR
dr_manager_ids     = np.repeat(manager_ids, team_sizes)
dr_treatment       = np.repeat(treatment, team_sizes)
dr_is_new_manager  = np.repeat(is_new_manager, team_sizes)

# DR demographics (random - not linked to manager)
dr_regions     = np.random.choice(REGIONS, size=total_dr)
dr_orgs        = np.random.choice(ORGANIZATIONS, size=total_dr)
dr_jobs        = np.random.choice(JOB_FAMILIES, size=total_dr)
dr_performance = np.random.choice(PERFORMANCE_RATINGS, size=total_dr, p=perf_probs)
dr_genders     = np.random.choice(GENDERS, size=total_dr, p=gender_probs)
dr_ages        = np.random.normal(35, 8, total_dr).clip(22, 65).round().astype(int)
dr_tenures     = np.random.gamma(4, 6, total_dr).clip(1, 120).round().astype(int)

# DR-level baselines (prior year) - independent draws
baseline_workload_dr = generate_baseline(3.2, 0.95, total_dr)
baseline_turnover_intention_dr = generate_baseline(2.6, 1.00, total_dr)

# baseline_manager_support_dr: used as baseline for manager_support outcome
# NaN if the manager is new (manager didn't have DRs last year)
baseline_manager_support_dr_raw = generate_baseline(3.4, 0.90, total_dr)
baseline_manager_support_dr = np.where(dr_is_new_manager == 1, np.nan,
                                       baseline_manager_support_dr_raw)

# ============================================================================
# SECTION 10: GENERATE DIRECT REPORT OUTCOMES (WITH CLUSTERING)
# ============================================================================

print("\nGenerating direct report outcomes with clustering...")

def generate_clustered_outcome_with_baseline(
        base_mean, base_sd, treatment_effect_d,
        treatment, manager_ids, baseline, baseline_r=0.45,
        icc=0.15):
    """
    Generate a clustered survey outcome (1-5) with manager random effects,
    a prior-year baseline predictor, and a treatment effect.
    """
    n_obs = len(treatment)
    unique_mgrs = np.unique(manager_ids)
    n_mgrs = len(unique_mgrs)

    # Variance components
    total_var   = base_sd ** 2
    between_var = total_var * icc
    within_var  = total_var * (1 - icc)

    # Manager random effects
    mgr_effects = np.random.normal(0, np.sqrt(between_var), n_mgrs)
    mgr_effect_map = dict(zip(unique_mgrs, mgr_effects))
    obs_mgr_effects = np.array([mgr_effect_map[m] for m in manager_ids])

    # Baseline component
    bl_filled = np.where(np.isnan(baseline), np.nanmean(baseline), baseline)
    bl_centered = bl_filled - np.nanmean(baseline)
    baseline_component = baseline_r * bl_centered

    # Treatment effect
    treat_effect = treatment_effect_d * base_sd * treatment

    # Individual residual
    residual_sd = np.sqrt(within_var * (1 - baseline_r ** 2))
    noise = np.random.normal(0, residual_sd, n_obs)

    scores = base_mean + baseline_component + treat_effect + obs_mgr_effects + noise
    return np.clip(np.round(scores, 1), SURVEY_MIN, SURVEY_MAX)


# Manager Support Index: d  0.25, significant, small
manager_support_dr = generate_clustered_outcome_with_baseline(
    base_mean=3.5, base_sd=0.95, treatment_effect_d=0.25,
    treatment=dr_treatment, manager_ids=dr_manager_ids,
    baseline=baseline_manager_support_dr_raw, baseline_r=0.45, icc=0.15
)
print("  [OK] manager_support_index  (d0.25, sig)")

# Workload Index (DR): d  0.25, significant, small
workload_dr = generate_clustered_outcome_with_baseline(
    base_mean=3.3, base_sd=1.00, treatment_effect_d=0.25,
    treatment=dr_treatment, manager_ids=dr_manager_ids,
    baseline=baseline_workload_dr, baseline_r=0.40, icc=0.12
)
print("  [OK] workload_index_dr      (d0.25, sig)")

# Turnover Intention Index (DR): d  0.01, non-significant
# (Training effect doesn't meaningfully cascade to DR turnover intentions)
turnover_intention_dr = generate_clustered_outcome_with_baseline(
    base_mean=2.7, base_sd=1.05, treatment_effect_d=0.01,
    treatment=dr_treatment, manager_ids=dr_manager_ids,
    baseline=baseline_turnover_intention_dr, baseline_r=0.45, icc=0.10
)
print("  [OK] turnover_intention_dr  (d0.01, ns)")

# Assemble DR dataframe
df_direct_reports = pd.DataFrame({
    'direct_report_id': np.arange(1, total_dr + 1),
    'manager_id': dr_manager_ids,
    'treatment': dr_treatment,
    'region': dr_regions,
    'organization': dr_orgs,
    'job_family': dr_jobs,
    'performance_rating': dr_performance,
    'gender': dr_genders,
    'age': dr_ages,
    'tenure_months': dr_tenures,
    'baseline_workload_dr': baseline_workload_dr,
    'baseline_turnover_intention_dr': baseline_turnover_intention_dr,
    'manager_support_index': manager_support_dr,
    'workload_index_dr': workload_dr,
    'turnover_intention_index_dr': turnover_intention_dr,
})

print(f"\n[OK] Direct report dataframe assembled: {df_direct_reports.shape}")

# ============================================================================
# SECTION 11: VERIFICATION - SELECTION PATTERNS
# ============================================================================

print("\n" + "=" * 80)
print("VERIFICATION: SELECTION PATTERNS (COVARIATE IMBALANCE)")
print("=" * 80)

# Treatment rates by organisation
print("\nTreatment rate by Organisation:")
org_rates = df_managers.groupby('organization')['treatment'].agg(['sum', 'count', 'mean'])
org_rates.columns = ['n_treated', 'n_total', 'rate']
org_rates = org_rates.sort_values('rate', ascending=False)
for org, row in org_rates.iterrows():
    bar = '-' * int(row['rate'] * 40)
    print(f"  {org:20s}  {row['rate']:.1%}  ({int(row['n_treated']):>3}/{int(row['n_total']):>4})  {bar}")

# Treatment rates by performance rating
print("\nTreatment rate by Performance Rating:")
perf_rates = df_managers.groupby('performance_rating')['treatment'].agg(['sum', 'count', 'mean'])
perf_rates.columns = ['n_treated', 'n_total', 'rate']
perf_rates = perf_rates.reindex(PERFORMANCE_RATINGS)
for rating, row in perf_rates.iterrows():
    bar = '-' * int(row['rate'] * 40)
    print(f"  {rating:20s}  {row['rate']:.1%}  ({int(row['n_treated']):>3}/{int(row['n_total']):>4})  {bar}")

# Treatment rates by new-manager status
print("\nTreatment rate by new-manager status:")
for flag, label in [(0, 'Experienced'), (1, 'New Manager')]:
    subset = df_managers[df_managers['is_new_manager'] == flag]
    rate = subset['treatment'].mean()
    print(f"  {label:20s}  {rate:.1%}  ({subset['treatment'].sum()}/{len(subset)})")

# Baseline means by treatment
print("\nBaseline means by treatment group:")
bl_vars = ['baseline_manager_efficacy', 'baseline_workload',
           'baseline_turnover_intention']
for var in bl_vars:
    t_mean = df_managers[df_managers['treatment'] == 1][var].mean()
    c_mean = df_managers[df_managers['treatment'] == 0][var].mean()
    diff = t_mean - c_mean
    print(f"  {var:35s}  Treated: {t_mean:.2f}  Control: {c_mean:.2f}  ={diff:+.2f}")

# Standardised mean differences (SMDs) for all covariates
print("\nStandardised Mean Differences (SMDs):")

def smd_continuous(series_t, series_c):
    """Compute absolute SMD for a continuous variable."""
    pooled_sd = np.sqrt((series_t.std()**2 + series_c.std()**2) / 2)
    if pooled_sd == 0:
        return 0.0
    return (series_t.mean() - series_c.mean()) / pooled_sd

t_df = df_managers[df_managers['treatment'] == 1]
c_df = df_managers[df_managers['treatment'] == 0]

smd_results = []
for var in ['age', 'tenure_months', 'team_size'] + bl_vars:
    s = smd_continuous(t_df[var].dropna(), c_df[var].dropna())
    flag_str = '  *** IMBALANCED' if abs(s) > 0.10 else ''
    print(f"  {var:35s}  SMD = {s:+.3f}{flag_str}")
    smd_results.append({'Variable': var, 'SMD': round(s, 3)})

# Categorical SMDs (difference in proportions / pooled SD)
for var in ['organization', 'region', 'performance_rating', 'gender']:
    dummies = pd.get_dummies(df_managers[var])
    for col in dummies.columns:
        s = smd_continuous(dummies.loc[treatment == 1, col], dummies.loc[treatment == 0, col])
        flag_str = '  *** IMBALANCED' if abs(s) > 0.10 else ''
        print(f"  {var}={col:25s}  SMD = {s:+.3f}{flag_str}")
        smd_results.append({'Variable': f'{var}={col}', 'SMD': round(s, 3)})

# Propensity score distribution
print("\nPropensity Score Distribution:")
for label, mask in [('Treated', treatment == 1), ('Control', treatment == 0)]:
    ps = propensity_scores[mask]
    print(f"  {label:10s}  mean={ps.mean():.3f}  sd={ps.std():.3f}  "
          f"min={ps.min():.3f}  max={ps.max():.3f}  "
          f"Q25={np.percentile(ps, 25):.3f}  Q75={np.percentile(ps, 75):.3f}")

# ============================================================================
# SECTION 12: VERIFICATION - DESCRIPTIVE STATISTICS
# ============================================================================

print("\n" + "=" * 80)
print("DESCRIPTIVE STATISTICS")
print("=" * 80)

print("\n--- MANAGER OUTCOMES ---")
print("\nRetention Rates by Treatment Group:")
for outcome in ['retention_3month', 'retention_6month', 'retention_9month', 'retention_12month']:
    t_rate = df_managers[df_managers['treatment'] == 1][outcome].mean()
    c_rate = df_managers[df_managers['treatment'] == 0][outcome].mean()
    print(f"  {outcome}: Treated = {t_rate:.1%}, Control = {c_rate:.1%}, Diff = {(t_rate-c_rate)*100:+.1f} pp")

print("\nSurvey Outcomes by Treatment Group (Mean +/- SD):")
survey_outcomes_mgr = ['manager_efficacy_index', 'workload_index_mgr', 'turnover_intention_index_mgr']
for outcome in survey_outcomes_mgr:
    t = df_managers[df_managers['treatment'] == 1][outcome]
    c = df_managers[df_managers['treatment'] == 0][outcome]
    print(f"  {outcome}:")
    print(f"    Treated:  {t.mean():.2f}  {t.std():.2f}")
    print(f"    Control:  {c.mean():.2f}  {c.std():.2f}")

print("\n--- DIRECT REPORT OUTCOMES ---")
survey_outcomes_dr = ['manager_support_index', 'workload_index_dr', 'turnover_intention_index_dr']
for outcome in survey_outcomes_dr:
    t = df_direct_reports[df_direct_reports['treatment'] == 1][outcome]
    c = df_direct_reports[df_direct_reports['treatment'] == 0][outcome]
    print(f"  {outcome}:")
    print(f"    Treated:  {t.mean():.2f}  {t.std():.2f}")
    print(f"    Control:  {c.mean():.2f}  {c.std():.2f}")

# ============================================================================
# SECTION 13: VERIFICATION - STATISTICAL TESTS
# ============================================================================

print("\n" + "=" * 80)
print("STATISTICAL TESTS - CONFIRMING EXPECTED RESULTS")
print("=" * 80)

print("\n--- MANAGER-LEVEL SURVEY OUTCOMES (t-tests) ---")
for outcome in survey_outcomes_mgr:
    t = df_managers[df_managers['treatment'] == 1][outcome]
    c = df_managers[df_managers['treatment'] == 0][outcome]
    t_stat, p_val = stats.ttest_ind(t, c)
    d = smd_continuous(t, c)
    sig = '***' if p_val < 0.001 else '**' if p_val < 0.01 else '*' if p_val < 0.05 else 'ns'
    print(f"\n  {outcome}:")
    print(f"    Cohen's d = {d:.3f}")
    print(f"    t = {t_stat:.2f}, p = {p_val:.4f}  {sig}")

print("\n--- MANAGER RETENTION (Logistic Regression) ---")
from sklearn.linear_model import LogisticRegression
for outcome in ['retention_3month', 'retention_6month', 'retention_9month', 'retention_12month']:
    X = df_managers[['treatment']].values
    y = df_managers[outcome].values
    model = LogisticRegression(max_iter=1000)
    model.fit(X, y)
    or_val = np.exp(model.coef_[0][0])
    print(f"  {outcome}:  OR = {or_val:.2f}")

print("\n--- DIRECT REPORT OUTCOMES (t-tests, ignoring clustering) ---")
print("  Note: proper analysis requires GEE to account for clustering.\n")
for outcome in survey_outcomes_dr:
    t = df_direct_reports[df_direct_reports['treatment'] == 1][outcome]
    c = df_direct_reports[df_direct_reports['treatment'] == 0][outcome]
    t_stat, p_val = stats.ttest_ind(t, c)
    d = smd_continuous(t, c)
    sig = '***' if p_val < 0.001 else '**' if p_val < 0.01 else '*' if p_val < 0.05 else 'ns'
    print(f"  {outcome}:")
    print(f"    Cohen's d = {d:.3f},  t = {t_stat:.2f},  p = {p_val:.4f}  {sig}")

# R&D heterogeneity check (Manager Efficacy)
print("\n--- R&D HETEROGENEITY CHECK (Manager Efficacy) ---")
rnd_t = df_managers[(df_managers['treatment'] == 1) & (df_managers['organization'] == 'R&D')]['manager_efficacy_index']
rnd_c = df_managers[(df_managers['treatment'] == 0) & (df_managers['organization'] == 'R&D')]['manager_efficacy_index']
oth_t = df_managers[(df_managers['treatment'] == 1) & (df_managers['organization'] != 'R&D')]['manager_efficacy_index']
oth_c = df_managers[(df_managers['treatment'] == 0) & (df_managers['organization'] != 'R&D')]['manager_efficacy_index']
d_rnd = smd_continuous(rnd_t, rnd_c)
d_oth = smd_continuous(oth_t, oth_c)
print(f"  R&D managers      :  d = {d_rnd:.3f}  (expected ~0.65)")
print(f"  Non-R&D managers  :  d = {d_oth:.3f}  (expected ~0.50)")

# New-manager heterogeneity check (Turnover Intention + Retention)
print("\n--- NEW-MANAGER HETEROGENEITY CHECK ---")
print("\n  Turnover Intention Index:")
new_t = df_managers[(df_managers['treatment'] == 1) & (df_managers['is_new_manager'] == 1)]['turnover_intention_index_mgr']
new_c = df_managers[(df_managers['treatment'] == 0) & (df_managers['is_new_manager'] == 1)]['turnover_intention_index_mgr']
exp_t = df_managers[(df_managers['treatment'] == 1) & (df_managers['is_new_manager'] == 0)]['turnover_intention_index_mgr']
exp_c = df_managers[(df_managers['treatment'] == 0) & (df_managers['is_new_manager'] == 0)]['turnover_intention_index_mgr']
d_new = smd_continuous(new_t, new_c)
d_exp = smd_continuous(exp_t, exp_c)
print(f"    New managers        :  d = {d_new:.3f}  (expected ~0.45)")
print(f"    Experienced managers:  d = {d_exp:.3f}  (expected ~0.25)")

print("\n  3-Month Retention Rate:")
for flag, label in [(1, 'New managers'), (0, 'Experienced')]:
    sub = df_managers[df_managers['is_new_manager'] == flag]
    tr = sub[sub['treatment'] == 1]['retention_3month'].mean()
    cr = sub[sub['treatment'] == 0]['retention_3month'].mean()
    print(f"    {label:22s}  Treated: {tr:.1%}  Control: {cr:.1%}  Diff = {(tr-cr)*100:+.1f} pp")

print("\n  12-Month Retention Rate:")
for flag, label in [(1, 'New managers'), (0, 'Experienced')]:
    sub = df_managers[df_managers['is_new_manager'] == flag]
    tr = sub[sub['treatment'] == 1]['retention_12month'].mean()
    cr = sub[sub['treatment'] == 0]['retention_12month'].mean()
    print(f"    {label:22s}  Treated: {tr:.1%}  Control: {cr:.1%}  Diff = {(tr-cr)*100:+.1f} pp")

# ============================================================================
# SECTION 14: EXPORT DATA
# ============================================================================

print("\n" + "=" * 80)
print("EXPORTING DATA")
print("=" * 80)

df_managers.to_csv('./data/s2_manager_data.csv', index=False)
df_direct_reports.to_csv('./data/s2_direct_report_data.csv', index=False)

print("\n[OK] Data exported to:")
print("  - ./data/s2_manager_data.csv")
print("  - ./data/s2_direct_report_data.csv")

# ============================================================================
# SECTION 15: EXCEL REPORT
# ============================================================================

print("\n" + "=" * 80)
print("GENERATING EXCEL DESCRIPTIVES REPORT")
print("=" * 80)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Formatting constants (same as S1)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
THIN_BORDER = Border(bottom=Side(style='thin', color='B0B0B0'))

GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def apply_header_format(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def apply_alternating_rows(ws, start_row, end_row, max_col, bold_keywords=None):
    if bold_keywords is None:
        bold_keywords = ['Overall', 'Total']
    for r in range(start_row, end_row + 1):
        fill = ALT_ROW_FILL if (r - start_row) % 2 == 0 else WHITE_FILL
        is_bold = False
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value is not None and any(kw in str(cell.value) for kw in bold_keywords):
                is_bold = True
        if is_bold:
            for col in range(1, max_col + 1):
                ws.cell(row=r, column=col).font = BOLD_FONT


def auto_fit_columns(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def write_title(ws, row, title_text, max_col=1):
    cell = ws.cell(row=row, column=1, value=title_text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left')
    if max_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)


def write_df(ws, df, start_row, start_col=1):
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=c_idx, value=str(col_name))
    apply_header_format(ws, start_row, start_col + len(df.columns) - 1)

    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)
            if isinstance(value, (np.integer,)):
                cell.value = int(value)
            elif isinstance(value, (np.floating,)):
                cell.value = round(float(value), 3)
            elif pd.isna(value):
                cell.value = ''
            else:
                cell.value = value

    end_row = start_row + len(df)
    max_col = start_col + len(df.columns) - 1
    apply_alternating_rows(ws, start_row + 1, end_row, max_col)
    return end_row


def apply_pvalue_conditional(ws, col_letter, start_row, end_row):
    for r in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{r}"]
        try:
            val = float(cell.value)
        except (TypeError, ValueError):
            continue
        if val < 0.05:
            cell.fill = GREEN_FILL
        elif val < 0.10:
            cell.fill = YELLOW_FILL
        else:
            cell.fill = RED_FILL


def descriptives_by_group(df, variables, group_col='treatment',
                          group_labels={1: 'Treated', 0: 'Control'}):
    rows = []
    for var in variables:
        for val, label in list(group_labels.items()) + [('all', 'Overall')]:
            s = df[var].dropna() if val == 'all' else df[df[group_col] == val][var].dropna()
            rows.append({
                'Variable': var, 'Group': label, 'n': len(s),
                'Mean': round(s.mean(), 3), 'SD': round(s.std(), 3),
                'Min': round(s.min(), 3), 'Max': round(s.max(), 3),
                'Median': round(s.median(), 3),
                'Skewness': round(s.skew(), 3),
                'Kurtosis': round(s.kurtosis(), 3),
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Create workbook
# ---------------------------------------------------------------------------
EXCEL_PATH = './data/s2_data_descriptives.xlsx'
wb = Workbook()

# ===== SHEET 1: README =====================================================
ws_readme = wb.active
ws_readme.title = "README"

n_treated_actual = int(treatment.sum())
n_control_actual = int(N_TOTAL - treatment.sum())
readme_lines = [
    ("Leadership Development Program - Scenario 2 Descriptive Report", TITLE_FONT),
    ("", None),
    ("Dataset Overview", BOLD_FONT),
    ("This workbook contains descriptive statistics and selection-pattern diagnostics for an", None),
    ("observational (open-enrollment) evaluation of a leadership development program.", None),
    ("", None),
    ("Design", BOLD_FONT),
    (f" Treated group: {n_treated_actual} managers who voluntarily enrolled (Jan-Mar)", None),
    (f" Control pool : {n_control_actual} managers who did not participate", None),
    (f" Total direct reports: {total_dr} (team sizes {MIN_TEAM_SIZE}-{MAX_TEAM_SIZE})", None),
    (f" ~{int(is_new_manager.sum())} managers flagged as new (no prior manager-level baselines)", None),
    (" Self-selection driven by: Organisation, performance rating", None),
    (" Seed for reproducibility: 42", None),
    ("", None),
    ("Key Difference vs Scenario 1", BOLD_FONT),
    (" Treatment is NOT randomised - covariates are imbalanced", None),
    (" Prior-year baseline scores included as controls for outcomes", None),
    (" New managers have NaN for baseline_manager_efficacy", None),
    (" No Below / Far Below performers are treated", None),
    ("", None),
    ("Sheets in This Workbook", BOLD_FONT),
    (" README - This summary sheet", None),
    (" Selection_Patterns - Treatment rates by organisation, performance, new-mgr status", None),
    (" Covariate_Balance - SMDs for all covariates (pre-weighting)", None),
    (" PS_Summary - Propensity score distribution statistics", None),
    (" Manager_Descriptives - Descriptive stats by treatment group", None),
    (" DR_Descriptives - Direct report descriptive stats by treatment group", None),
    (" Retention_Summary - Retention rates at 3, 6, 9, 12 months with chi-square tests", None),
    (" Period_Turnover_Rates - Period-specific turnover rates", None),
    (" Manager_Outcomes - Manager survey outcomes with Cohen's d and t-test p-values", None),
    (" DR_Outcomes - Direct report survey outcomes with Cohen's d and t-test p-values", None),
    (" Team_Size_Distribution - Team size distribution statistics", None),
    (" Raw_Managers - Full manager-level dataset", None),
    (" Raw_Direct_Reports - Full direct-report-level dataset", None),
    ("", None),
    ("Formatting Key", BOLD_FONT),
    (" p-value cells: Green (p < .05), Yellow (.05 <= p < .10), Red (p >= .10)", None),
    (" Bold rows indicate 'Overall' or 'Total' summaries", None),
]
for r_idx, (text, font) in enumerate(readme_lines, start=1):
    cell = ws_readme.cell(row=r_idx, column=1, value=text)
    if font:
        cell.font = font
ws_readme.column_dimensions['A'].width = 90
print("  [OK] README")

# ===== SHEET 2: Selection_Patterns ==========================================
ws_sel = wb.create_sheet("Selection_Patterns")
write_title(ws_sel, 1, "Treatment Selection Patterns (Self-Selection Drivers)", max_col=5)

# By Organisation
sel_org_rows = []
for org in sorted(ORGANIZATIONS):
    sub = df_managers[df_managers['organization'] == org]
    n_t = sub['treatment'].sum()
    n_total = len(sub)
    sel_org_rows.append({
        'Organisation': org,
        'n Treated': int(n_t),
        'n Total': int(n_total),
        'Participation Rate %': round(n_t / n_total * 100, 1) if n_total > 0 else 0
    })
sel_org_rows.append({
    'Organisation': 'Overall',
    'n Treated': int(treatment.sum()),
    'n Total': int(N_TOTAL),
    'Participation Rate %': round(treatment.mean() * 100, 1)
})
df_sel_org = pd.DataFrame(sel_org_rows)

row_ptr = 3
write_title(ws_sel, row_ptr, "By Organisation", max_col=4)
row_ptr = write_df(ws_sel, df_sel_org, start_row=row_ptr + 2) + 2

# By New-Manager Status
sel_new_rows = []
for flag, label in [(0, 'Experienced'), (1, 'New Manager')]:
    sub = df_managers[df_managers['is_new_manager'] == flag]
    n_t = sub['treatment'].sum()
    n_total = len(sub)
    sel_new_rows.append({
        'Status': label, 'n Treated': int(n_t), 'n Total': int(n_total),
        'Participation Rate %': round(n_t / n_total * 100, 1)
    })
df_sel_new = pd.DataFrame(sel_new_rows)
write_title(ws_sel, row_ptr, "By New-Manager Status", max_col=4)
row_ptr = write_df(ws_sel, df_sel_new, start_row=row_ptr + 2) + 2

# By Performance Rating
sel_perf_rows = []
for rating in PERFORMANCE_RATINGS:
    sub = df_managers[df_managers['performance_rating'] == rating]
    n_t = sub['treatment'].sum()
    n_total = len(sub)
    sel_perf_rows.append({
        'Performance Rating': rating,
        'n Treated': int(n_t),
        'n Total': int(n_total),
        'Participation Rate %': round(n_t / n_total * 100, 1) if n_total > 0 else 0
    })
df_sel_perf = pd.DataFrame(sel_perf_rows)
write_title(ws_sel, row_ptr, "By Performance Rating", max_col=4)
row_ptr = write_df(ws_sel, df_sel_perf, start_row=row_ptr + 2)

auto_fit_columns(ws_sel)
print("  [OK] Selection_Patterns")

# ===== SHEET 3: Covariate_Balance ===========================================
ws_cov = wb.create_sheet("Covariate_Balance")
write_title(ws_cov, 1, "Covariate Balance  Standardised Mean Differences (Pre-Weighting)", max_col=6)

cov_bal_rows = []
# Continuous
for var in ['age', 'tenure_months', 'team_size', 'baseline_manager_efficacy', 'baseline_workload',
            'baseline_turnover_intention']:
    tv = t_df[var].dropna()
    cv = c_df[var].dropna()
    s = smd_continuous(tv, cv)
    t_stat, p_val = stats.ttest_ind(tv, cv, equal_var=False)
    cov_bal_rows.append({
        'Variable': var,
        'Treated Mean': round(tv.mean(), 3),
        'Control Mean': round(cv.mean(), 3),
        'SMD': round(s, 3),
        't-statistic': round(t_stat, 3),
        'p-value': round(p_val, 4),
    })

# Categorical - proportion difference as SMD
for var in ['organization', 'region', 'performance_rating', 'gender', 'is_new_manager']:
    if var == 'is_new_manager':
        dummies = pd.DataFrame({var: df_managers[var]})
    else:
        dummies = pd.get_dummies(df_managers[var], prefix=var)
    for col in dummies.columns:
        tv = dummies.loc[treatment == 1, col].astype(float)
        cv = dummies.loc[treatment == 0, col].astype(float)
        s = smd_continuous(tv, cv)
        t_stat, p_val = stats.ttest_ind(tv, cv, equal_var=False)
        cov_bal_rows.append({
            'Variable': col,
            'Treated Mean': round(tv.mean(), 3),
            'Control Mean': round(cv.mean(), 3),
            'SMD': round(s, 3),
            't-statistic': round(t_stat, 3),
            'p-value': round(p_val, 4),
        })

df_cov_bal = pd.DataFrame(cov_bal_rows)
end_r = write_df(ws_cov, df_cov_bal, start_row=3)
pval_col = get_column_letter(6)
apply_pvalue_conditional(ws_cov, pval_col, 4, end_r)
auto_fit_columns(ws_cov)
print("  [OK] Covariate_Balance")

# ===== SHEET 4: PS_Summary =================================================
ws_ps = wb.create_sheet("PS_Summary")
write_title(ws_ps, 1, "Propensity Score Distribution (Pre-Weighting)", max_col=8)

ps_rows = []
for label, mask in [('Treated', treatment == 1), ('Control', treatment == 0), ('Overall', np.ones(N_TOTAL, dtype=bool))]:
    ps = propensity_scores[mask]
    ps_rows.append({
        'Group': label, 'n': int(mask.sum()),
        'Mean': round(ps.mean(), 4), 'SD': round(ps.std(), 4),
        'Min': round(ps.min(), 4), 'Q25': round(np.percentile(ps, 25), 4),
        'Median': round(np.median(ps), 4), 'Q75': round(np.percentile(ps, 75), 4),
        'Max': round(ps.max(), 4),
    })
df_ps = pd.DataFrame(ps_rows)
end_r = write_df(ws_ps, df_ps, start_row=3)
auto_fit_columns(ws_ps)
print("  [OK] PS_Summary")

# ===== SHEET 5: Manager_Descriptives ========================================
ws_mgr_desc = wb.create_sheet("Manager_Descriptives")
mgr_cont_vars = ['age', 'tenure_months', 'team_size',
                  'baseline_manager_efficacy', 'baseline_workload',
                  'baseline_turnover_intention',
                  'manager_efficacy_index', 'workload_index_mgr', 'turnover_intention_index_mgr']
df_mgr_desc = descriptives_by_group(df_managers, mgr_cont_vars)
write_title(ws_mgr_desc, 1, "Manager-Level Descriptive Statistics (Continuous & Likert Variables)", max_col=len(df_mgr_desc.columns))
write_df(ws_mgr_desc, df_mgr_desc, start_row=3)
auto_fit_columns(ws_mgr_desc)
print("  [OK] Manager_Descriptives")

# ===== SHEET 6: DR_Descriptives =============================================
ws_dr_desc = wb.create_sheet("DR_Descriptives")
dr_cont_vars = ['age', 'tenure_months',
                 'baseline_workload_dr', 'baseline_turnover_intention_dr',
                 'manager_support_index', 'workload_index_dr', 'turnover_intention_index_dr']
df_dr_desc = descriptives_by_group(df_direct_reports, dr_cont_vars)
write_title(ws_dr_desc, 1, "Direct Report-Level Descriptive Statistics", max_col=len(df_dr_desc.columns))
write_df(ws_dr_desc, df_dr_desc, start_row=3)
auto_fit_columns(ws_dr_desc)
print("  [OK] DR_Descriptives")

# ===== SHEET 7: Retention_Summary ===========================================
ws_ret = wb.create_sheet("Retention_Summary")
write_title(ws_ret, 1, "Retention Rates by Treatment Group with Chi-Square Tests", max_col=10)

ret_rows = []
for outcome, label in [('retention_3month', '3-Month'), ('retention_6month', '6-Month'),
                        ('retention_9month', '9-Month'), ('retention_12month', '12-Month')]:
    t_vals = df_managers[df_managers['treatment'] == 1][outcome]
    c_vals = df_managers[df_managers['treatment'] == 0][outcome]
    ct = pd.crosstab(df_managers['treatment'], df_managers[outcome])
    chi2, p_val, _, _ = stats.chi2_contingency(ct)
    ret_rows.append({
        'Timepoint': label,
        'Treated Retained': int(t_vals.sum()),
        'Treated Lost': int(len(t_vals) - t_vals.sum()),
        'Treated Rate %': round(t_vals.mean() * 100, 1),
        'Control Retained': int(c_vals.sum()),
        'Control Lost': int(len(c_vals) - c_vals.sum()),
        'Control Rate %': round(c_vals.mean() * 100, 1),
        'Difference (pp)': round((t_vals.mean() - c_vals.mean()) * 100, 1),
        'Chi-square': round(chi2, 3),
        'p-value': round(p_val, 4),
    })
df_ret = pd.DataFrame(ret_rows)
end_r = write_df(ws_ret, df_ret, start_row=3)
pval_col = get_column_letter(10)
apply_pvalue_conditional(ws_ret, pval_col, 4, end_r)
auto_fit_columns(ws_ret)
print("  [OK] Retention_Summary")

# ===== SHEET 8: Period_Turnover_Rates ========================================
ws_period = wb.create_sheet("Period_Turnover_Rates")
write_title(ws_period, 1, "Period-Specific Turnover Rates Analysis", max_col=10)

t_mgr = df_managers[df_managers['treatment'] == 1]
c_mgr = df_managers[df_managers['treatment'] == 0]
n_t = len(t_mgr)
n_c = len(c_mgr)

period_data = []
prev_t_retained, prev_c_retained = n_t, n_c
for i, (col, lbl) in enumerate([
    ('retention_3month', '0-3 Months'), ('retention_6month', '3-6 Months'),
    ('retention_9month', '6-9 Months'),  ('retention_12month', '9-12 Months')]):

    curr_t = int(t_mgr[col].sum())
    curr_c = int(c_mgr[col].sum())
    lost_t = prev_t_retained - curr_t
    lost_c = prev_c_retained - curr_c
    rate_t = lost_t / prev_t_retained * 100 if prev_t_retained > 0 else 0
    rate_c = lost_c / prev_c_retained * 100 if prev_c_retained > 0 else 0

    contingency = [[curr_t, lost_t], [curr_c, lost_c]]
    try:
        chi2, p_val, _, _ = stats.chi2_contingency(contingency)
    except ValueError:
        chi2, p_val = 0, 1.0

    period_data.append({
        'Period': lbl,
        'Treated At Risk': prev_t_retained, 'Treated Lost': lost_t,
        'Treated Turnover %': round(rate_t, 1),
        'Control At Risk': prev_c_retained, 'Control Lost': lost_c,
        'Control Turnover %': round(rate_c, 1),
        'Difference (pp)': round(rate_c - rate_t, 1),
        'Chi-square': round(chi2, 3), 'p-value': round(p_val, 4)
    })
    prev_t_retained = curr_t
    prev_c_retained = curr_c

df_period = pd.DataFrame(period_data)
end_r = write_df(ws_period, df_period, start_row=3)
pval_col = get_column_letter(10)
apply_pvalue_conditional(ws_period, pval_col, 4, end_r)
auto_fit_columns(ws_period)
print("  [OK] Period_Turnover_Rates")

# ===== SHEET 9: Manager_Outcomes ============================================
ws_mgr_out = wb.create_sheet("Manager_Outcomes")
write_title(ws_mgr_out, 1, "Manager-Level Survey Outcomes  Treatment vs Control", max_col=10)

mgr_out_rows = []
for outcome in survey_outcomes_mgr:
    t_vals = df_managers[df_managers['treatment'] == 1][outcome]
    c_vals = df_managers[df_managers['treatment'] == 0][outcome]
    t_stat, p_val = stats.ttest_ind(t_vals, c_vals)
    d = smd_continuous(t_vals, c_vals)
    mgr_out_rows.append({
        'Outcome': outcome,
        'Treated n': len(t_vals), 'Treated Mean': round(t_vals.mean(), 3),
        'Treated SD': round(t_vals.std(), 3),
        'Control n': len(c_vals), 'Control Mean': round(c_vals.mean(), 3),
        'Control SD': round(c_vals.std(), 3),
        "Cohen's d": round(d, 3),
        't-statistic': round(t_stat, 3), 'p-value': round(p_val, 4),
    })
df_mgr_out = pd.DataFrame(mgr_out_rows)
end_r = write_df(ws_mgr_out, df_mgr_out, start_row=3)
pval_col = get_column_letter(10)
apply_pvalue_conditional(ws_mgr_out, pval_col, 4, end_r)
auto_fit_columns(ws_mgr_out)
print("  [OK] Manager_Outcomes")

# ===== SHEET 10: DR_Outcomes ================================================
ws_dr_out = wb.create_sheet("DR_Outcomes")
write_title(ws_dr_out, 1, "Direct Report-Level Survey Outcomes  Treatment vs Control", max_col=10)

dr_out_rows = []
for outcome in survey_outcomes_dr:
    t_vals = df_direct_reports[df_direct_reports['treatment'] == 1][outcome]
    c_vals = df_direct_reports[df_direct_reports['treatment'] == 0][outcome]
    t_stat, p_val = stats.ttest_ind(t_vals, c_vals)
    d = smd_continuous(t_vals, c_vals)
    dr_out_rows.append({
        'Outcome': outcome,
        'Treated n': len(t_vals), 'Treated Mean': round(t_vals.mean(), 3),
        'Treated SD': round(t_vals.std(), 3),
        'Control n': len(c_vals), 'Control Mean': round(c_vals.mean(), 3),
        'Control SD': round(c_vals.std(), 3),
        "Cohen's d": round(d, 3),
        't-statistic': round(t_stat, 3), 'p-value': round(p_val, 4),
    })
df_dr_out = pd.DataFrame(dr_out_rows)
end_r = write_df(ws_dr_out, df_dr_out, start_row=3)
pval_col = get_column_letter(10)
apply_pvalue_conditional(ws_dr_out, pval_col, 4, end_r)
auto_fit_columns(ws_dr_out)
print("  [OK] DR_Outcomes")

# ===== SHEET 11: Team_Size_Distribution =====================================
ws_team = wb.create_sheet("Team_Size_Distribution")
write_title(ws_team, 1, "Team Size Distribution Across Managers", max_col=5)

team_s = pd.Series(team_sizes)
team_summary = pd.DataFrame([
    {'Statistic': 'Min',                   'Value': int(team_s.min())},
    {'Statistic': 'Max',                   'Value': int(team_s.max())},
    {'Statistic': 'Mean',                  'Value': round(team_s.mean(), 2)},
    {'Statistic': 'SD',                    'Value': round(team_s.std(), 2)},
    {'Statistic': 'Median',                'Value': round(team_s.median(), 1)},
    {'Statistic': 'Total Direct Reports',  'Value': int(team_s.sum())},
    {'Statistic': 'Total Managers',        'Value': int(len(team_s))},
])
end_r = write_df(ws_team, team_summary, start_row=3)

freq = team_s.value_counts().sort_index()
freq_df = pd.DataFrame({
    'Team Size': freq.index.astype(int),
    'Count': freq.values.astype(int),
    '% of Managers': (freq.values / len(team_s) * 100).round(1),
    'Cumulative %': (freq.values.cumsum() / len(team_s) * 100).round(1),
})
total_row = pd.DataFrame([{
    'Team Size': 'Total', 'Count': int(freq.values.sum()),
    '% of Managers': 100.0, 'Cumulative %': 100.0
}])
freq_df = pd.concat([freq_df, total_row], ignore_index=True)

write_title(ws_team, end_r + 2, "Frequency Table", max_col=4)
write_df(ws_team, freq_df, start_row=end_r + 4)
auto_fit_columns(ws_team)
print("  [OK] Team_Size_Distribution")

# ===== SHEET 12: Raw_Managers ===============================================
ws_raw_mgr = wb.create_sheet("Raw_Managers")
write_title(ws_raw_mgr, 1, "Full Manager-Level Dataset", max_col=len(df_managers.columns))
write_df(ws_raw_mgr, df_managers, start_row=3)
auto_fit_columns(ws_raw_mgr)
print("  [OK] Raw_Managers")

# ===== SHEET 13: Raw_Direct_Reports =========================================
ws_raw_dr = wb.create_sheet("Raw_Direct_Reports")
write_title(ws_raw_dr, 1, "Full Direct Report-Level Dataset", max_col=len(df_direct_reports.columns))
write_df(ws_raw_dr, df_direct_reports, start_row=3)
auto_fit_columns(ws_raw_dr)
print("  [OK] Raw_Direct_Reports")

# ---------------------------------------------------------------------------
# Save workbook
# ---------------------------------------------------------------------------
wb.save(EXCEL_PATH)
print(f"\n[DONE] Excel report saved to: {EXCEL_PATH}")

print("\n" + "=" * 80)
print("DATA GENERATION COMPLETE")
print("=" * 80)
print(f"\nFiles created:")
print(f"  ./data/s2_manager_data.csv        ({df_managers.shape[0]} rows x {df_managers.shape[1]} cols)")
print(f"  ./data/s2_direct_report_data.csv  ({df_direct_reports.shape[0]} rows x {df_direct_reports.shape[1]} cols)")
print(f"  ./data/s2_data_descriptives.xlsx  (13 sheets)")

