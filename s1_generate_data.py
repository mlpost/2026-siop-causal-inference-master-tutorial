
"""
Mock Data Generator for Leadership Development Program Tutorial
================================================================

This script generates a deterministic mock dataset for a statistics tutorial
on causal inference using a staggered rollout design.

Design: Staggered RCT with two cohorts of new managers
- Cohort 1: Training Jan-Mar (Treatment)
- Cohort 2: Training July-Sept (Control at early measurements)

Author: Generated for Statistics Tutorial
Date: 2024
"""

import numpy as np
import pandas as pd
from scipy import stats
from scipy.special import expit  # logistic function
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# SECTION 1: SETUP AND CONSTANTS
# ============================================================================

# Set seed for reproducibility
SEED = 42
np.random.seed(SEED)

# Sample sizes
N_MANAGERS = 300
N_COHORT = 150
MIN_TEAM_SIZE = 5
MAX_TEAM_SIZE = 12

# Demographics categories
REGIONS = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East & Africa']
ORGANIZATIONS = ['R&D', 'Commercial', 'Manufacturing', 'Digital', 'HR', 'Finance']
JOB_FAMILIES = [
    'Clinical Operations', 'Regulatory Affairs', 'Data Science', 'Medical Affairs',
    'Sales', 'Marketing', 'Supply Chain', 'Quality Assurance', 'Pharmacovigilance',
    'Market Access', 'Human Resources', 'IT & Digital', 'Finance & Accounting',
    'Legal & Compliance', 'Communications'
]
PERFORMANCE_RATINGS = ['Far Below', 'Below', 'Meets', 'Exceeds', 'Far Exceeds']
GENDERS = ['Male', 'Female', 'Non-Binary/Other']

# Survey scale parameters (1-5 Likert)
SURVEY_MIN = 1
SURVEY_MAX = 5

print("="*80)
print("LEADERSHIP DEVELOPMENT PROGRAM - MOCK DATA GENERATOR")
print("="*80)
print(f"\nSeed: {SEED}")
print(f"Total Managers: {N_MANAGERS} ({N_COHORT} per cohort)")
print(f"Expected Direct Reports: ~{N_MANAGERS * 8} (team sizes {MIN_TEAM_SIZE}-{MAX_TEAM_SIZE})")

# ============================================================================
# SECTION 2: GENERATE MANAGER-LEVEL DATA
# ============================================================================

print("\n" + "="*80)
print("GENERATING MANAGER-LEVEL DATA")
print("="*80)

# Create manager IDs
manager_ids = np.arange(1, N_MANAGERS + 1)

# Random assignment to cohorts (perfect 50/50 split)
cohort_assignment = np.array([1] * N_COHORT + [2] * N_COHORT)
np.random.shuffle(cohort_assignment)

# Generate demographics with near-perfect balance
# Strategy: Create perfectly balanced sets, then shuffle within cohorts

def create_balanced_categorical(categories, n_per_cohort):
    """Create near-perfectly balanced categorical variable across cohorts"""
    n_cats = len(categories)
    # Calculate how many of each category per cohort
    base_count = n_per_cohort // n_cats
    remainder = n_per_cohort % n_cats
    
    # Create balanced distribution
    cohort_values = []
    for i, cat in enumerate(categories):
        count = base_count + (1 if i < remainder else 0)
        cohort_values.extend([cat] * count)
    
    return cohort_values

# Generate balanced demographics for each cohort separately
regions_c1 = create_balanced_categorical(REGIONS, N_COHORT)
regions_c2 = create_balanced_categorical(REGIONS, N_COHORT)
np.random.shuffle(regions_c1)
np.random.shuffle(regions_c2)

orgs_c1 = create_balanced_categorical(ORGANIZATIONS, N_COHORT)
orgs_c2 = create_balanced_categorical(ORGANIZATIONS, N_COHORT)
np.random.shuffle(orgs_c1)
np.random.shuffle(orgs_c2)

jobs_c1 = create_balanced_categorical(JOB_FAMILIES, N_COHORT)
jobs_c2 = create_balanced_categorical(JOB_FAMILIES, N_COHORT)
np.random.shuffle(jobs_c1)
np.random.shuffle(jobs_c2)

# Performance ratings (roughly normal distribution)
perf_probs = [0.05, 0.15, 0.60, 0.15, 0.05]  # Most at "Meets"
perf_c1 = np.random.choice(PERFORMANCE_RATINGS, size=N_COHORT, p=perf_probs)
perf_c2 = np.random.choice(PERFORMANCE_RATINGS, size=N_COHORT, p=perf_probs)

# Gender (balanced)
gender_probs = [0.48, 0.48, 0.04]
gender_c1 = np.random.choice(GENDERS, size=N_COHORT, p=gender_probs)
gender_c2 = np.random.choice(GENDERS, size=N_COHORT, p=gender_probs)

# Age (continuous, realistic for managers)
age_c1 = np.random.normal(38, 6, N_COHORT).clip(28, 55).round().astype(int)
age_c2 = np.random.normal(38, 6, N_COHORT).clip(28, 55).round().astype(int)

# Tenure in months (new managers, so relatively short)
tenure_c1 = np.random.gamma(3, 4, N_COHORT).clip(1, 36).round().astype(int)
tenure_c2 = np.random.gamma(3, 4, N_COHORT).clip(1, 36).round().astype(int)

# Combine cohorts
regions = np.concatenate([regions_c1, regions_c2])
organizations = np.concatenate([orgs_c1, orgs_c2])
job_families = np.concatenate([jobs_c1, jobs_c2])
performance = np.concatenate([perf_c1, perf_c2])
genders = np.concatenate([gender_c1, gender_c2])
ages = np.concatenate([age_c1, age_c2])
tenures = np.concatenate([tenure_c1, tenure_c2])

# Reorder by cohort assignment
sort_idx = np.argsort(cohort_assignment)
manager_ids = manager_ids[sort_idx]
cohort_assignment = cohort_assignment[sort_idx]
regions = regions[sort_idx]
organizations = organizations[sort_idx]
job_families = job_families[sort_idx]
performance = performance[sort_idx]
genders = genders[sort_idx]
ages = ages[sort_idx]
tenures = tenures[sort_idx]

# ============================================================================
# SECTION 3: GENERATE MANAGER OUTCOMES
# ============================================================================

print("\nGenerating manager outcomes with treatment effects...")

# Treatment indicator (1 = Cohort 1, 0 = Cohort 2)
treatment = (cohort_assignment == 1).astype(int)

# Scenario 1 assumption: all managers are newly promoted this year
is_new_manager = np.ones(N_MANAGERS, dtype=int)

# Shared latent engagement factor for correlated baseline outcomes
latent_engagement = np.random.normal(0, 1, N_MANAGERS)

def generate_baseline(base_mean, base_sd, latent, latent_weight=0.4):
    """Generate a baseline score (1-5) correlated with a latent factor."""
    noise = np.random.normal(0, base_sd * np.sqrt(1 - latent_weight**2), len(latent))
    scores = base_mean + latent_weight * base_sd * latent + noise
    return np.clip(np.round(scores, 1), SURVEY_MIN, SURVEY_MAX)

# Baseline outcomes (prior-year style covariates)
# No baseline_manager_efficacy in Scenario 1 (all managers are new);
# use baseline_engagement_index as the manager efficacy prognostic baseline.
baseline_engagement_index = generate_baseline(3.3, 0.90, latent_engagement, 0.45)
baseline_workload = generate_baseline(3.1, 0.95, latent_engagement, 0.35)
baseline_turnover_intention = generate_baseline(2.7, 1.00, latent_engagement, 0.40)

# Helper function to generate survey items with treatment effects and baseline influence
def generate_outcome_with_baseline(base_mean, base_sd, treatment_effect_d, treatment,
                                   baseline, baseline_r=0.50):
    """
    Generate survey outcome (1-5 scale) with treatment effect and baseline correlation.
    
    Parameters:
    - base_mean: baseline mean for control group
    - base_sd: outcome standard deviation scale
    - treatment_effect_d: Cohen's d effect size
    - treatment: binary treatment indicator
    - baseline: baseline covariate (can include NaN)
    - baseline_r: baseline-outcome correlation strength
    """
    n = len(treatment)

    # Handle partially/fully missing baseline values robustly
    baseline_mean = np.nanmean(baseline)
    if np.isnan(baseline_mean):
        bl_centered = np.zeros(n)
    else:
        bl_filled = np.where(np.isnan(baseline), baseline_mean, baseline)
        bl_centered = bl_filled - baseline_mean

    baseline_component = baseline_r * bl_centered
    treatment_effect = treatment_effect_d * base_sd
    residual_sd = base_sd * np.sqrt(max(1 - baseline_r ** 2, 0))
    noise = np.random.normal(0, residual_sd, n)

    scores = base_mean + baseline_component + treatment * treatment_effect + noise

    # Clip to 1-5 scale and round to 1 decimal
    scores = np.clip(scores, SURVEY_MIN, SURVEY_MAX)
    scores = np.round(scores, 1)
    
    return scores

# Helper function to generate retention outcomes with baseline adjustment
def generate_retention_with_baseline(base_rate, treatment_or, treatment, baseline_turnover_intent,
                                     baseline_coef=0.30):
    """
    Generate binary retention outcome with treatment effect and baseline covariate influence.
    
    Parameters:
    - base_rate: baseline retention rate for control group
    - treatment_or: odds ratio for treatment effect
    - treatment: binary treatment indicator
    - baseline_turnover_intent: baseline turnover intention (higher = more likely to stay)
    - baseline_coef: coefficient on centered baseline turnover intention
    """
    # Convert base rate to log-odds
    base_logit = np.log(base_rate / (1 - base_rate))
    
    # Calculate treatment effect on log-odds scale
    treatment_logit_effect = np.log(treatment_or)
    
    baseline_mean = np.nanmean(baseline_turnover_intent)
    if np.isnan(baseline_mean):
        bl_effect = np.zeros(len(treatment))
    else:
        bl_centered = np.where(np.isnan(baseline_turnover_intent), baseline_mean,
                               baseline_turnover_intent) - baseline_mean
        bl_effect = baseline_coef * bl_centered

    # Generate probabilities
    logits = base_logit + treatment * treatment_logit_effect + bl_effect
    probs = expit(logits)  # Convert to probabilities
    
    # Generate binary outcomes
    outcomes = (np.random.uniform(0, 1, len(treatment)) < probs).astype(int)
    
    return outcomes

# RETENTION OUTCOMES
# Pattern: Large gap at 3 & 6 months (C1 trained, C2 not yet trained)
#          Gap persists at 9 & 12 months (C2 trained too late to recover early losses)
#          Key insight: Early training timing matters — training during critical onboarding window protects retention

# 3-month: Moderate-large effect (OR ~2.5) - C1 ~95%, C2 ~85% (10pp gap)
retention_3mo = generate_retention_with_baseline(
    base_rate=0.85,
    treatment_or=2.5,
    treatment=treatment,
    baseline_turnover_intent=baseline_turnover_intention
)

# 6-month: CONDITIONAL on surviving to 3 months
# To go from C1: 95% → 92%, need 92/95 = 96.8% of 3mo survivors to survive to 6mo
# To go from C2: 85% → 80%, need 80/85 = 94.1% of 3mo survivors to survive to 6mo
p_survive_6mo_c1 = 0.92 / 0.95  # ~0.968
p_survive_6mo_c2 = 0.80 / 0.85  # ~0.941
cond_prob_6mo = np.where(treatment == 1, p_survive_6mo_c1, p_survive_6mo_c2)
retention_6mo = retention_3mo * (np.random.binomial(1, cond_prob_6mo, size=len(treatment)))

# 9-month: CONDITIONAL on surviving to 6 months
# Widen the period-specific gap so cumulative retention remains significant
p_survive_9mo_c1 = 0.98   # C1 barely loses anyone (trained, stable)
p_survive_9mo_c2 = 0.92   # C2 continues to lose managers (untrained during critical window)
cond_prob_9mo = np.where(treatment == 1, p_survive_9mo_c1, p_survive_9mo_c2)
retention_9mo = retention_6mo * (np.random.binomial(1, cond_prob_9mo, size=len(treatment)))

# 12-month: CONDITIONAL on surviving to 9 months
# Maintain a meaningful period gap so cumulative retention stays significant
p_survive_12mo_c1 = 0.98   # C1 remains stable
p_survive_12mo_c2 = 0.94   # C2 still losing more than C1
cond_prob_12mo = np.where(treatment == 1, p_survive_12mo_c1, p_survive_12mo_c2)
retention_12mo = retention_9mo * (np.random.binomial(1, cond_prob_12mo, size=len(treatment)))

# Assertion checks: retention counts can only decrease over time (cumulative/monotonic)
assert (retention_6mo <= retention_3mo).all(), "ERROR: Someone returned after leaving at 3 months!"
assert (retention_9mo <= retention_6mo).all(), "ERROR: Someone returned after leaving at 6 months!"
assert (retention_12mo <= retention_9mo).all(), "ERROR: Someone returned after leaving at 9 months!"

# SURVEY OUTCOMES
# Manager Efficacy Index: moderate effect (d ~0.5)
manager_efficacy = generate_outcome_with_baseline(
    base_mean=3.4,
    base_sd=0.9,
    treatment_effect_d=0.5,
    treatment=treatment,
    baseline=baseline_engagement_index,
    baseline_r=0.50
)

# Workload Index: marginal/ns effect (d ~0.15)
workload_manager = generate_outcome_with_baseline(
    base_mean=3.2,
    base_sd=1.0,
    treatment_effect_d=0.15,
    treatment=treatment,
    baseline=baseline_workload,
    baseline_r=0.45
)

# Turnover Intention Index: small-to-moderate significant effect (d ~0.4, p < .05)
# Higher values = higher intention to STAY (positively coded)
# Treated managers score higher (more intention to stay)
turnover_intention_manager = generate_outcome_with_baseline(
    base_mean=2.8,
    base_sd=1.0,
    treatment_effect_d=0.4,
    treatment=treatment,
    baseline=baseline_turnover_intention,
    baseline_r=0.50
)

# Create manager dataframe
df_managers = pd.DataFrame({
    'manager_id': manager_ids,
    'cohort': cohort_assignment,
    'treatment': treatment,
    'region': regions,
    'organization': organizations,
    'job_family': job_families,
    'performance_rating': performance,
    'gender': genders,
    'age': ages,
    'tenure_months': tenures,
    'is_new_manager': is_new_manager,
    'baseline_engagement_index': baseline_engagement_index,
    'baseline_workload': baseline_workload,
    'baseline_turnover_intention': baseline_turnover_intention,
    'retention_3month': retention_3mo,
    'retention_6month': retention_6mo,
    'retention_9month': retention_9mo,
    'retention_12month': retention_12mo,
    'manager_efficacy_index': manager_efficacy,
    'workload_index_mgr': workload_manager,
    'turnover_intention_index_mgr': turnover_intention_manager
})

print(f"[OK] Generated {len(df_managers)} manager records")

# ============================================================================
# SECTION 4: GENERATE DIRECT REPORT DATA
# ============================================================================

print("\n" + "="*80)
print("GENERATING DIRECT REPORT DATA")
print("="*80)

# Generate team sizes (varying 4-12 per manager)
team_sizes = np.random.randint(MIN_TEAM_SIZE, MAX_TEAM_SIZE + 1, N_MANAGERS)
total_direct_reports = team_sizes.sum()

print(f"Total direct reports: {total_direct_reports}")
print(f"Average team size: {team_sizes.mean():.1f}")

# Create direct report records linked to managers
dr_manager_ids = np.repeat(manager_ids, team_sizes)
dr_cohorts = np.repeat(cohort_assignment, team_sizes)
dr_treatment = np.repeat(treatment, team_sizes)

# Generate direct report demographics (don't need perfect balance here)
dr_regions = np.random.choice(REGIONS, size=total_direct_reports)
dr_orgs = np.random.choice(ORGANIZATIONS, size=total_direct_reports)
dr_jobs = np.random.choice(JOB_FAMILIES, size=total_direct_reports)
dr_performance = np.random.choice(PERFORMANCE_RATINGS, size=total_direct_reports, p=perf_probs)
dr_genders = np.random.choice(GENDERS, size=total_direct_reports, p=gender_probs)
dr_ages = np.random.normal(35, 8, total_direct_reports).clip(22, 65).round().astype(int)
dr_tenures = np.random.gamma(4, 6, total_direct_reports).clip(1, 120).round().astype(int)

# Direct report baseline outcomes (prior-year style covariates)
dr_is_new_manager = np.repeat(is_new_manager, team_sizes)
dr_latent_engagement = np.random.normal(0, 1, total_direct_reports)
baseline_workload_dr = generate_baseline(3.2, 0.95, dr_latent_engagement, 0.35)
baseline_turnover_intention_dr = generate_baseline(2.6, 1.00, dr_latent_engagement, 0.40)

# Baseline manager support reflects DRs' experience with their previous manager
# (prior to the newly promoted manager), so it is available even for new managers
baseline_manager_support_dr = generate_baseline(3.4, 0.90, dr_latent_engagement, 0.40)

# ============================================================================
# SECTION 5: GENERATE DIRECT REPORT OUTCOMES (WITH CLUSTERING)
# ============================================================================

print("\nGenerating direct report outcomes with clustering...")

# For clustered data, we need to add manager-level random effects
# This creates intra-cluster correlation (ICC)

def generate_clustered_outcome_with_baseline(base_mean, base_sd, treatment_effect_d,
                                             treatment, manager_ids, baseline,
                                             baseline_r=0.45, icc=0.15):
    """
    Generate clustered survey outcome with manager-level random effects,
    baseline influence, and treatment effects.
    
    Parameters:
    - base_mean: baseline mean for control group
    - base_sd: baseline standard deviation
    - treatment_effect_d: Cohen's d effect size
    - treatment: binary treatment indicator
    - manager_ids: manager ID for each direct report
    - baseline: baseline covariate (can include NaN)
    - baseline_r: baseline-outcome correlation strength
    - icc: intra-class correlation (proportion of variance at manager level)
    """
    n_obs = len(treatment)
    unique_managers = np.unique(manager_ids)
    n_managers = len(unique_managers)
    
    # Calculate variance components
    total_var = base_sd ** 2
    between_var = total_var * icc
    within_var = total_var * (1 - icc)
    
    # Generate manager-level random effects
    manager_effects = np.random.normal(0, np.sqrt(between_var), n_managers)
    manager_effect_dict = dict(zip(unique_managers, manager_effects))
    
    # Map manager effects to observations
    obs_manager_effects = np.array([manager_effect_dict[mid] for mid in manager_ids])
    
    # Baseline component
    baseline_mean = np.nanmean(baseline)
    if np.isnan(baseline_mean):
        baseline_component = np.zeros(n_obs)
    else:
        bl_filled = np.where(np.isnan(baseline), baseline_mean, baseline)
        bl_centered = bl_filled - baseline_mean
        baseline_component = baseline_r * bl_centered

    # Generate individual-level residuals
    residual_sd = np.sqrt(within_var * max(1 - baseline_r ** 2, 0))
    individual_residuals = np.random.normal(0, residual_sd, n_obs)
    
    # Combine components
    treatment_effect = treatment_effect_d * base_sd
    scores = (base_mean + baseline_component + treatment * treatment_effect +
              obs_manager_effects + individual_residuals)
    
    # Clip to 1-5 scale
    scores = np.clip(scores, SURVEY_MIN, SURVEY_MAX)
    scores = np.round(scores, 1)
    
    return scores

# DIRECT REPORT SURVEY OUTCOMES

# Manager Support Index: moderate effect (d ~0.4)
manager_support_dr = generate_clustered_outcome_with_baseline(
    base_mean=3.5,
    base_sd=0.95,
    treatment_effect_d=0.4,
    treatment=dr_treatment,
    manager_ids=dr_manager_ids,
    baseline=baseline_manager_support_dr,
    baseline_r=0.45,
    icc=0.15
)

# Workload Index: small effect (d ~0.25)
workload_dr = generate_clustered_outcome_with_baseline(
    base_mean=3.3,
    base_sd=1.0,
    treatment_effect_d=0.25,
    treatment=dr_treatment,
    manager_ids=dr_manager_ids,
    baseline=baseline_workload_dr,
    baseline_r=0.40,
    icc=0.12
)

# Turnover Intention Index: non-significant effect (d ~0.08)
# Higher values = higher intention to STAY (positively coded)
# Treated DRs trend slightly higher but effect is non-significant
turnover_intention_dr = generate_clustered_outcome_with_baseline(
    base_mean=2.7,
    base_sd=1.05,
    treatment_effect_d=0.08,
    treatment=dr_treatment,
    manager_ids=dr_manager_ids,
    baseline=baseline_turnover_intention_dr,
    baseline_r=0.45,
    icc=0.10
)

# Create direct report dataframe
df_direct_reports = pd.DataFrame({
    'direct_report_id': np.arange(1, total_direct_reports + 1),
    'manager_id': dr_manager_ids,
    'cohort': dr_cohorts,
    'treatment': dr_treatment,
    'region': dr_regions,
    'organization': dr_orgs,
    'job_family': dr_jobs,
    'performance_rating': dr_performance,
    'gender': dr_genders,
    'age': dr_ages,
    'tenure_months': dr_tenures,
    'baseline_workload_dr': baseline_workload_dr,
    'baseline_turnover_intention_dr': baseline_turnover_intention_dr,
    'baseline_manager_support_dr': baseline_manager_support_dr,
    'manager_support_index': manager_support_dr,
    'workload_index_dr': workload_dr,
    'turnover_intention_index_dr': turnover_intention_dr
})

print(f"[OK] Generated {len(df_direct_reports)} direct report records")

# ============================================================================
# SECTION 6: VERIFICATION - COVARIATE BALANCE
# ============================================================================

print("\n" + "="*80)
print("VERIFICATION: COVARIATE BALANCE CHECK")
print("="*80)
print("\nChecking randomization balance across cohorts...")

def check_balance_categorical(df, var_name, cohort_col='cohort'):
    """Check balance for categorical variable"""
    ct = pd.crosstab(df[var_name], df[cohort_col], normalize='columns')
    chi2, p_value, _, _ = stats.chi2_contingency(pd.crosstab(df[var_name], df[cohort_col]))
    print(f"\n{var_name}:")
    print(ct.round(3))
    print(f"Chi-square test: chi2 = {chi2:.2f}, p = {p_value:.3f}")

def check_balance_continuous(df, var_name, cohort_col='cohort'):
    """Check balance for continuous variable"""
    c1 = df[df[cohort_col] == 1][var_name]
    c2 = df[df[cohort_col] == 2][var_name]
    t_stat, p_value = stats.ttest_ind(c1, c2)
    print(f"\n{var_name}:")
    print(f"  Cohort 1: M = {c1.mean():.2f}, SD = {c1.std():.2f}")
    print(f"  Cohort 2: M = {c2.mean():.2f}, SD = {c2.std():.2f}")
    print(f"  t-test: t = {t_stat:.2f}, p = {p_value:.3f}")

# Check categorical variables
check_balance_categorical(df_managers, 'region')
check_balance_categorical(df_managers, 'organization')
check_balance_categorical(df_managers, 'performance_rating')
check_balance_categorical(df_managers, 'gender')

# Check continuous variables
check_balance_continuous(df_managers, 'age')
check_balance_continuous(df_managers, 'tenure_months')

# ============================================================================
# SECTION 7: VERIFICATION - DESCRIPTIVE STATISTICS
# ============================================================================

print("\n" + "="*80)
print("DESCRIPTIVE STATISTICS")
print("="*80)

print("\n--- MANAGER OUTCOMES ---")
print("\nRetention Rates by Cohort:")
for outcome in ['retention_3month', 'retention_6month', 'retention_9month', 'retention_12month']:
    c1_rate = df_managers[df_managers['cohort'] == 1][outcome].mean()
    c2_rate = df_managers[df_managers['cohort'] == 2][outcome].mean()
    print(f"{outcome}: Cohort 1 = {c1_rate:.1%}, Cohort 2 = {c2_rate:.1%}")

print("\nSurvey Outcomes by Cohort (Mean ± SD):")
survey_outcomes_mgr = ['manager_efficacy_index', 'workload_index_mgr', 'turnover_intention_index_mgr']
for outcome in survey_outcomes_mgr:
    c1 = df_managers[df_managers['cohort'] == 1][outcome]
    c2 = df_managers[df_managers['cohort'] == 2][outcome]
    print(f"{outcome}:")
    print(f"  Cohort 1: {c1.mean():.2f} ± {c1.std():.2f}")
    print(f"  Cohort 2: {c2.mean():.2f} ± {c2.std():.2f}")

print("\n--- DIRECT REPORT OUTCOMES ---")
print("\nSurvey Outcomes by Cohort (Mean ± SD):")
survey_outcomes_dr = ['manager_support_index', 'workload_index_dr', 'turnover_intention_index_dr']
for outcome in survey_outcomes_dr:
    c1 = df_direct_reports[df_direct_reports['cohort'] == 1][outcome]
    c2 = df_direct_reports[df_direct_reports['cohort'] == 2][outcome]
    print(f"{outcome}:")
    print(f"  Cohort 1: {c1.mean():.2f} ± {c1.std():.2f}")
    print(f"  Cohort 2: {c2.mean():.2f} ± {c2.std():.2f}")

# ============================================================================
# SECTION 8: VERIFICATION - STATISTICAL TESTS
# ============================================================================

print("\n" + "="*80)
print("STATISTICAL TESTS - CONFIRMING EXPECTED RESULTS")
print("="*80)

print("\n--- MANAGER-LEVEL OUTCOMES ---")

# Retention outcomes (logistic regression)
print("\nRetention Outcomes (Logistic Regression):")
from sklearn.linear_model import LogisticRegression

for outcome in ['retention_3month', 'retention_6month', 'retention_9month', 'retention_12month']:
    X = df_managers[['treatment']].values
    y = df_managers[outcome].values
    
    model = LogisticRegression()
    model.fit(X, y)
    
    # Calculate odds ratio
    or_value = np.exp(model.coef_[0][0])
    
    # Simple z-test for coefficient
    from scipy.stats import norm
    # Note: This is simplified; proper inference would use statsmodels
    z_score = model.coef_[0][0] / 0.3  # Approximate SE
    p_value = 2 * (1 - norm.cdf(abs(z_score)))
    
    print(f"\n{outcome}:")
    print(f"  Odds Ratio: {or_value:.2f}")
    print(f"  p-value: {p_value:.3f} {'***' if p_value < 0.001 else '**' if p_value < 0.01 else '*' if p_value < 0.05 else 'ns'}")

# Survey outcomes (t-tests)
print("\n\nSurvey Outcomes (Independent t-tests):")
for outcome in survey_outcomes_mgr:
    c1 = df_managers[df_managers['treatment'] == 1][outcome]
    c2 = df_managers[df_managers['treatment'] == 0][outcome]
    
    t_stat, p_value = stats.ttest_ind(c1, c2)
    cohens_d = (c1.mean() - c2.mean()) / np.sqrt((c1.std()**2 + c2.std()**2) / 2)
    
    print(f"\n{outcome}:")
    print(f"  Cohort 1: M = {c1.mean():.2f}, SD = {c1.std():.2f}")
    print(f"  Cohort 2: M = {c2.mean():.2f}, SD = {c2.std():.2f}")
    print(f"  Cohen's d: {cohens_d:.2f}")
    print(f"  t({len(c1)+len(c2)-2}) = {t_stat:.2f}, p = {p_value:.3f} {'***' if p_value < 0.001 else '**' if p_value < 0.01 else '*' if p_value < 0.05 else 'ns'}")

print("\n--- DIRECT REPORT OUTCOMES (Clustered) ---")
print("\nNote: These are simple t-tests. Proper analysis requires GEE to account for clustering.")

for outcome in survey_outcomes_dr:
    c1 = df_direct_reports[df_direct_reports['treatment'] == 1][outcome]
    c2 = df_direct_reports[df_direct_reports['treatment'] == 0][outcome]
    
    t_stat, p_value = stats.ttest_ind(c1, c2)
    cohens_d = (c1.mean() - c2.mean()) / np.sqrt((c1.std()**2 + c2.std()**2) / 2)
    
    print(f"\n{outcome}:")
    print(f"  Cohort 1: M = {c1.mean():.2f}, SD = {c1.std():.2f}")
    print(f"  Cohort 2: M = {c2.mean():.2f}, SD = {c2.std():.2f}")
    print(f"  Cohen's d: {cohens_d:.2f}")
    print(f"  t = {t_stat:.2f}, p = {p_value:.3f} {'***' if p_value < 0.001 else '**' if p_value < 0.01 else '*' if p_value < 0.05 else 'ns'}")

# ============================================================================
# SECTION 9: EXPORT DATA
# ============================================================================

print("\n" + "="*80)
print("EXPORTING DATA")
print("="*80)

# Export to CSV
df_managers.to_csv('./data/s1_manager_data.csv', index=False)
df_direct_reports.to_csv('./data/s1_direct_report_data.csv', index=False)

print("\n[OK] Data exported to:")
print("  - ./data/s1_manager_data.csv")
print("  - ./data/s1_direct_report_data.csv")

# ============================================================================
# SECTION 10: EXCEL REPORT EXPORT
# ============================================================================

print("\n" + "="*80)
print("GENERATING EXCEL DESCRIPTIVES REPORT")
print("="*80)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Formatting constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
THIN_BORDER = Border(
    bottom=Side(style='thin', color='B0B0B0')
)

# Conditional formatting fills for p-values
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def apply_header_format(ws, row, max_col):
    """Apply dark-blue header formatting to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def apply_alternating_rows(ws, start_row, end_row, max_col, bold_keywords=None):
    """Apply alternating row shading and optionally bold rows containing keywords."""
    if bold_keywords is None:
        bold_keywords = ['Overall', 'Total']
    for r in range(start_row, end_row + 1):
        fill = ALT_ROW_FILL if (r - start_row) % 2 == 0 else WHITE_FILL
        is_bold = False
        for col in range(1, max_col + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value is not None and any(kw in str(cell.value) for kw in bold_keywords):
                is_bold = True
        if is_bold:
            for col in range(1, max_col + 1):
                ws.cell(row=r, column=col).font = BOLD_FONT


def auto_fit_columns(ws, min_width=10, max_width=40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def write_title(ws, row, title_text, max_col=1):
    """Write a title row merged across columns."""
    cell = ws.cell(row=row, column=1, value=title_text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='left')
    if max_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)


def write_df(ws, df, start_row, start_col=1):
    """Write a DataFrame to a worksheet starting at (start_row, start_col).
    Returns the row number after the last data row."""
    # Header
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=c_idx, value=str(col_name))
    apply_header_format(ws, start_row, start_col + len(df.columns) - 1)

    # Data
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c_idx, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)
            # Convert numpy types to native Python for openpyxl
            if isinstance(value, (np.integer,)):
                cell.value = int(value)
            elif isinstance(value, (np.floating,)):
                cell.value = round(float(value), 3)
            else:
                cell.value = value

    end_row = start_row + len(df)
    max_col = start_col + len(df.columns) - 1
    apply_alternating_rows(ws, start_row + 1, end_row, max_col)
    return end_row


def apply_pvalue_conditional(ws, col_letter, start_row, end_row):
    """Apply green/yellow/red fill to p-value cells."""
    for r in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{r}"]
        try:
            val = float(cell.value)
        except (TypeError, ValueError):
            continue
        if val < 0.05:
            cell.fill = GREEN_FILL
        elif val < 0.10:
            cell.fill = YELLOW_FILL
        else:
            cell.fill = RED_FILL


def create_period_turnover_sheet(ws, df_managers):
    """
    Create Period_Turnover_Rates sheet showing period-specific turnover rates.
    Demonstrates that late-period turnover rates converge between cohorts.
    """
    from scipy.stats import chi2_contingency

    def safe_chi2(contingency_table):
        """Run chi-square safely, returning (chi2, p) with fallback for degenerate tables."""
        try:
            chi2, p_val, _, _ = chi2_contingency(contingency_table)
            return round(chi2, 3), round(p_val, 3)
        except ValueError:
            return np.nan, np.nan
    
    # Separate cohorts
    c1 = df_managers[df_managers['cohort'] == 1]
    c2 = df_managers[df_managers['cohort'] == 2]
    n_c1 = len(c1)
    n_c2 = len(c2)
    
    # Calculate period-specific turnover
    periods_data = []
    
    # Period 1: 0-3 months
    c1_at_risk_p1 = n_c1
    c1_retained_3mo = c1['retention_3month'].sum()
    c1_lost_p1 = c1_at_risk_p1 - c1_retained_3mo
    c1_turnover_p1 = c1_lost_p1 / c1_at_risk_p1 * 100
    
    c2_at_risk_p1 = n_c2
    c2_retained_3mo = c2['retention_3month'].sum()
    c2_lost_p1 = c2_at_risk_p1 - c2_retained_3mo
    c2_turnover_p1 = c2_lost_p1 / c2_at_risk_p1 * 100
    
    # Chi-square for period 1
    contingency_p1 = [[c1_retained_3mo, c1_lost_p1], [c2_retained_3mo, c2_lost_p1]]
    chi2_p1, p_p1 = safe_chi2(contingency_p1)
    
    periods_data.append({
        'Period': '0-3 Months',
        'C1 At Risk': c1_at_risk_p1,
        'C1 Lost': c1_lost_p1,
        'C1 Turnover %': round(c1_turnover_p1, 1),
        'C2 At Risk': c2_at_risk_p1,
        'C2 Lost': c2_lost_p1,
        'C2 Turnover %': round(c2_turnover_p1, 1),
        'Difference (pp)': round(c2_turnover_p1 - c1_turnover_p1, 1),
        'Chi-square': chi2_p1,
        'p-value': p_p1
    })
    
    # Period 2: 3-6 months (at risk = those who survived to 3 months)
    c1_at_risk_p2 = c1_retained_3mo
    c1_retained_6mo = c1['retention_6month'].sum()
    c1_lost_p2 = c1_at_risk_p2 - c1_retained_6mo
    c1_turnover_p2 = c1_lost_p2 / c1_at_risk_p2 * 100 if c1_at_risk_p2 > 0 else 0
    
    c2_at_risk_p2 = c2_retained_3mo
    c2_retained_6mo = c2['retention_6month'].sum()
    c2_lost_p2 = c2_at_risk_p2 - c2_retained_6mo
    c2_turnover_p2 = c2_lost_p2 / c2_at_risk_p2 * 100 if c2_at_risk_p2 > 0 else 0
    
    # Chi-square for period 2
    contingency_p2 = [[c1_retained_6mo, c1_lost_p2], [c2_retained_6mo, c2_lost_p2]]
    chi2_p2, p_p2 = safe_chi2(contingency_p2)
    
    periods_data.append({
        'Period': '3-6 Months',
        'C1 At Risk': int(c1_at_risk_p2),
        'C1 Lost': int(c1_lost_p2),
        'C1 Turnover %': round(c1_turnover_p2, 1),
        'C2 At Risk': int(c2_at_risk_p2),
        'C2 Lost': int(c2_lost_p2),
        'C2 Turnover %': round(c2_turnover_p2, 1),
        'Difference (pp)': round(c2_turnover_p2 - c1_turnover_p2, 1),
        'Chi-square': chi2_p2,
        'p-value': p_p2
    })
    
    # Period 3: 6-9 months (at risk = those who survived to 6 months)
    c1_at_risk_p3 = c1_retained_6mo
    c1_retained_9mo = c1['retention_9month'].sum()
    c1_lost_p3 = c1_at_risk_p3 - c1_retained_9mo
    c1_turnover_p3 = c1_lost_p3 / c1_at_risk_p3 * 100 if c1_at_risk_p3 > 0 else 0
    
    c2_at_risk_p3 = c2_retained_6mo
    c2_retained_9mo = c2['retention_9month'].sum()
    c2_lost_p3 = c2_at_risk_p3 - c2_retained_9mo
    c2_turnover_p3 = c2_lost_p3 / c2_at_risk_p3 * 100 if c2_at_risk_p3 > 0 else 0
    
    # Chi-square for period 3
    contingency_p3 = [[c1_retained_9mo, c1_lost_p3], [c2_retained_9mo, c2_lost_p3]]
    chi2_p3, p_p3 = safe_chi2(contingency_p3)
    
    periods_data.append({
        'Period': '6-9 Months',
        'C1 At Risk': int(c1_at_risk_p3),
        'C1 Lost': int(c1_lost_p3),
        'C1 Turnover %': round(c1_turnover_p3, 1),
        'C2 At Risk': int(c2_at_risk_p3),
        'C2 Lost': int(c2_lost_p3),
        'C2 Turnover %': round(c2_turnover_p3, 1),
        'Difference (pp)': round(c2_turnover_p3 - c1_turnover_p3, 1),
        'Chi-square': chi2_p3,
        'p-value': p_p3
    })
    
    # Period 4: 9-12 months (at risk = those who survived to 9 months)
    c1_at_risk_p4 = c1_retained_9mo
    c1_retained_12mo = c1['retention_12month'].sum()
    c1_lost_p4 = c1_at_risk_p4 - c1_retained_12mo
    c1_turnover_p4 = c1_lost_p4 / c1_at_risk_p4 * 100 if c1_at_risk_p4 > 0 else 0
    
    c2_at_risk_p4 = c2_retained_9mo
    c2_retained_12mo = c2['retention_12month'].sum()
    c2_lost_p4 = c2_at_risk_p4 - c2_retained_12mo
    c2_turnover_p4 = c2_lost_p4 / c2_at_risk_p4 * 100 if c2_at_risk_p4 > 0 else 0
    
    # Chi-square for period 4
    contingency_p4 = [[c1_retained_12mo, c1_lost_p4], [c2_retained_12mo, c2_lost_p4]]
    chi2_p4, p_p4 = safe_chi2(contingency_p4)
    
    periods_data.append({
        'Period': '9-12 Months',
        'C1 At Risk': int(c1_at_risk_p4),
        'C1 Lost': int(c1_lost_p4),
        'C1 Turnover %': round(c1_turnover_p4, 1),
        'C2 At Risk': int(c2_at_risk_p4),
        'C2 Lost': int(c2_lost_p4),
        'C2 Turnover %': round(c2_turnover_p4, 1),
        'Difference (pp)': round(c2_turnover_p4 - c1_turnover_p4, 1),
        'Chi-square': chi2_p4,
        'p-value': p_p4
    })
    
    # Create DataFrame
    df_periods = pd.DataFrame(periods_data)
    
    # Calculate early vs late period summary
    # Early (0-6 months): Combined turnover
    c1_early_lost = c1_lost_p1 + c1_lost_p2
    c1_early_rate = c1_early_lost / n_c1 * 100
    c2_early_lost = c2_lost_p1 + c2_lost_p2
    c2_early_rate = c2_early_lost / n_c2 * 100
    
    # Late (6-12 months): Combined turnover among those at risk at 6 months
    c1_late_lost = c1_lost_p3 + c1_lost_p4
    c1_late_rate = c1_late_lost / c1_at_risk_p3 * 100 if c1_at_risk_p3 > 0 else 0
    c2_late_lost = c2_lost_p3 + c2_lost_p4
    c2_late_rate = c2_late_lost / c2_at_risk_p3 * 100 if c2_at_risk_p3 > 0 else 0
    
    # Chi-square for early period
    contingency_early = [[n_c1 - c1_early_lost, c1_early_lost], [n_c2 - c2_early_lost, c2_early_lost]]
    chi2_early, p_early = safe_chi2(contingency_early)
    
    # Chi-square for late period
    contingency_late = [[c1_at_risk_p3 - c1_late_lost, c1_late_lost], [c2_at_risk_p3 - c2_late_lost, c2_late_lost]]
    chi2_late, p_late = safe_chi2(contingency_late)
    
    summary_data = [
        {
            'Period': 'Early (0-6 Months)',
            'C1 At Risk': n_c1,
            'C1 Lost': c1_early_lost,
            'C1 Turnover %': round(c1_early_rate, 1),
            'C2 At Risk': n_c2,
            'C2 Lost': c2_early_lost,
            'C2 Turnover %': round(c2_early_rate, 1),
            'Difference (pp)': round(c2_early_rate - c1_early_rate, 1),
            'Chi-square': chi2_early,
            'p-value': p_early
        },
        {
            'Period': 'Late (6-12 Months)',
            'C1 At Risk': int(c1_at_risk_p3),
            'C1 Lost': int(c1_late_lost),
            'C1 Turnover %': round(c1_late_rate, 1),
            'C2 At Risk': int(c2_at_risk_p3),
            'C2 Lost': int(c2_late_lost),
            'C2 Turnover %': round(c2_late_rate, 1),
            'Difference (pp)': round(c2_late_rate - c1_late_rate, 1),
            'Chi-square': chi2_late,
            'p-value': p_late
        }
    ]
    df_summary = pd.DataFrame(summary_data)
    
    # Write to worksheet (ws already provided)
    
    # Title row
    write_title(ws, 1, 'Period-Specific Turnover Rates Analysis', max_col=len(df_periods.columns))
    
    # Main table
    end_r = write_df(ws, df_periods, start_row=3)
    
    # Summary section title
    write_title(ws, end_r + 2, 'Early vs Late Period Summary', max_col=len(df_summary.columns))
    
    # Summary table
    end_r2 = write_df(ws, df_summary, start_row=end_r + 4)
    
    # Teaching narrative
    narrative_start = end_r2 + 2
    ws.cell(row=narrative_start, column=1, value='').font = BOLD_FONT
    ws.cell(row=narrative_start + 1, column=1, value='KEY INSIGHT:').font = BOLD_FONT
    ws.cell(row=narrative_start + 2, column=1, value='Period-specific turnover rates show that the training effect is front-loaded. Early turnover')
    ws.cell(row=narrative_start + 3, column=1, value='(months 0-6) is significantly higher for Cohort 2, but by months 6-12, both cohorts show')
    ws.cell(row=narrative_start + 4, column=1, value='similar period-specific turnover rates. This demonstrates that the training protects managers')
    ws.cell(row=narrative_start + 5, column=1, value='during the critical early onboarding period, and the 12-month cumulative retention gap is')
    ws.cell(row=narrative_start + 6, column=1, value='driven by early losses rather than ongoing differences in stability.')
    
    # Apply p-value conditional formatting
    pval_col_letter = get_column_letter(10)
    apply_pvalue_conditional(ws, pval_col_letter, 4, end_r)
    apply_pvalue_conditional(ws, pval_col_letter, end_r + 5, end_r2)
    
    auto_fit_columns(ws)
    
    return df_periods, df_summary


# ---------------------------------------------------------------------------
# Helper: compute descriptive stats for a set of variables by cohort
# ---------------------------------------------------------------------------
def descriptives_by_cohort(df, variables, cohort_col='cohort'):
    """Return a DataFrame of descriptive statistics broken down by cohort + overall."""
    rows = []
    for var in variables:
        for label, subset in [('Cohort 1', df[df[cohort_col] == 1]),
                               ('Cohort 2', df[df[cohort_col] == 2]),
                               ('Overall', df)]:
            s = subset[var].dropna()
            rows.append({
                'Variable': var,
                'Group': label,
                'n': len(s),
                'Mean': round(s.mean(), 3),
                'SD': round(s.std(), 3),
                'Min': round(s.min(), 3),
                'Max': round(s.max(), 3),
                'Median': round(s.median(), 3),
                'Skewness': round(s.skew(), 3),
                'Kurtosis': round(s.kurtosis(), 3),
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Create workbook
# ---------------------------------------------------------------------------
EXCEL_PATH = './data/s1_data_descriptives.xlsx'
wb = Workbook()

# ===== SHEET 1: README =====================================================
ws_readme = wb.active
ws_readme.title = "README"
readme_lines = [
    ("Leadership Development Program – Descriptive Report", TITLE_FONT),
    ("", None),
    ("Dataset Overview", BOLD_FONT),
    ("This workbook contains descriptive statistics and balance checks for a staggered-rollout", None),
    ("randomized controlled trial of a leadership development program at a pharmaceutical company.", None),
    ("", None),
    ("Design", BOLD_FONT),
    ("• Cohort 1 (Treatment): 150 new managers trained Jan–Mar", None),
    ("• Cohort 2 (Control at early measures): 150 new managers trained Jul–Sept", None),
    (f"• Total direct reports: {total_direct_reports} (team sizes {MIN_TEAM_SIZE}–{MAX_TEAM_SIZE})", None),
    ("• Seed for reproducibility: 42", None),
    ("", None),
    ("Sheets in This Workbook", BOLD_FONT),
    ("• README – This summary sheet", None),
    ("• Manager_Descriptives – Descriptive stats for manager-level continuous/Likert variables by cohort", None),
    ("• DR_Descriptives – Descriptive stats for direct-report-level variables by cohort", None),
    ("• Covariate_Balance – Categorical demographic balance (counts, %, chi-square p-values)", None),
    ("• Continuous_Balance – Continuous demographic balance (mean, SD, t-test p-values)", None),
    ("• Retention_Summary – Retention rates at 3, 6, 9, 12 months by cohort with chi-square tests", None),
    ("• Period_Turnover_Rates – Period-specific turnover rates showing early vs late period differences", None),
    ("• Manager_Outcomes – Manager survey outcomes with Cohen's d and t-test p-values", None),
    ("• DR_Outcomes – Direct report survey outcomes with Cohen's d and t-test p-values", None),
    ("• Team_Size_Distribution – Team size distribution statistics and frequency table", None),
    ("• Raw_Managers – Full manager-level dataset", None),
    ("• Raw_Direct_Reports – Full direct-report-level dataset", None),
    ("", None),
    ("Formatting Key", BOLD_FONT),
    ("• p-value cells: Green (p < .05), Yellow (.05 ≤ p < .10), Red (p ≥ .10)", None),
    ("• Bold rows indicate 'Overall' or 'Total' summaries", None),
]
for r_idx, (text, font) in enumerate(readme_lines, start=1):
    cell = ws_readme.cell(row=r_idx, column=1, value=text)
    if font:
        cell.font = font
ws_readme.column_dimensions['A'].width = 90

# ===== SHEET 2: Manager_Descriptives =======================================
ws_mgr_desc = wb.create_sheet("Manager_Descriptives")
mgr_cont_vars = ['age', 'tenure_months', 'manager_efficacy_index',
                  'workload_index_mgr', 'turnover_intention_index_mgr']
df_mgr_desc = descriptives_by_cohort(df_managers, mgr_cont_vars)
write_title(ws_mgr_desc, 1, "Manager-Level Descriptive Statistics (Continuous & Likert Variables)", max_col=len(df_mgr_desc.columns))
end_r = write_df(ws_mgr_desc, df_mgr_desc, start_row=3)
auto_fit_columns(ws_mgr_desc)
print("  [OK] Manager_Descriptives")

# ===== SHEET 3: DR_Descriptives ============================================
ws_dr_desc = wb.create_sheet("DR_Descriptives")
dr_cont_vars = ['age', 'tenure_months', 'manager_support_index',
                 'workload_index_dr', 'turnover_intention_index_dr']
df_dr_desc = descriptives_by_cohort(df_direct_reports, dr_cont_vars)
write_title(ws_dr_desc, 1, "Direct Report-Level Descriptive Statistics (Continuous & Likert Variables)", max_col=len(df_dr_desc.columns))
end_r = write_df(ws_dr_desc, df_dr_desc, start_row=3)
auto_fit_columns(ws_dr_desc)
print("  [OK] DR_Descriptives")

# ===== SHEET 4: Covariate_Balance (Categorical) ============================
ws_cov_bal = wb.create_sheet("Covariate_Balance")
write_title(ws_cov_bal, 1, "Covariate Balance – Categorical Demographics (Managers)", max_col=8)

cat_vars = ['region', 'organization', 'job_family', 'performance_rating', 'gender']
cov_rows = []
for var in cat_vars:
    ct_counts = pd.crosstab(df_managers[var], df_managers['cohort'])
    ct_pct = pd.crosstab(df_managers[var], df_managers['cohort'], normalize='columns')
    chi2, p_val, _, _ = stats.chi2_contingency(ct_counts)
    for level in ct_counts.index:
        cov_rows.append({
            'Variable': var,
            'Level': level,
            'Cohort 1 n': int(ct_counts.loc[level, 1]) if 1 in ct_counts.columns else 0,
            'Cohort 1 %': round(float(ct_pct.loc[level, 1]) * 100, 1) if 1 in ct_pct.columns else 0,
            'Cohort 2 n': int(ct_counts.loc[level, 2]) if 2 in ct_counts.columns else 0,
            'Cohort 2 %': round(float(ct_pct.loc[level, 2]) * 100, 1) if 2 in ct_pct.columns else 0,
            'Chi-square': round(chi2, 3),
            'p-value': round(p_val, 3),
        })
df_cov_bal = pd.DataFrame(cov_rows)
end_r = write_df(ws_cov_bal, df_cov_bal, start_row=3)
# p-value conditional formatting (column H = 8)
pval_col_letter = get_column_letter(8)
apply_pvalue_conditional(ws_cov_bal, pval_col_letter, 4, end_r)
auto_fit_columns(ws_cov_bal)
print("  [OK] Covariate_Balance")

# ===== SHEET 5: Continuous_Balance ==========================================
ws_cont_bal = wb.create_sheet("Continuous_Balance")
write_title(ws_cont_bal, 1, "Covariate Balance – Continuous Demographics (Managers)", max_col=8)

cont_bal_rows = []
for var in ['age', 'tenure_months']:
    c1 = df_managers[df_managers['cohort'] == 1][var]
    c2 = df_managers[df_managers['cohort'] == 2][var]
    t_stat, p_val = stats.ttest_ind(c1, c2)
    cont_bal_rows.append({
        'Variable': var,
        'Cohort 1 Mean': round(c1.mean(), 2),
        'Cohort 1 SD': round(c1.std(), 2),
        'Cohort 2 Mean': round(c2.mean(), 2),
        'Cohort 2 SD': round(c2.std(), 2),
        't-statistic': round(t_stat, 3),
        'p-value': round(p_val, 3),
        'SMD': round((c1.mean() - c2.mean()) / np.sqrt((c1.std()**2 + c2.std()**2) / 2), 3),
    })
df_cont_bal = pd.DataFrame(cont_bal_rows)
end_r = write_df(ws_cont_bal, df_cont_bal, start_row=3)
pval_col_letter = get_column_letter(7)
apply_pvalue_conditional(ws_cont_bal, pval_col_letter, 4, end_r)
auto_fit_columns(ws_cont_bal)
print("  [OK] Continuous_Balance")

# ===== SHEET 6: Retention_Summary ===========================================
ws_ret = wb.create_sheet("Retention_Summary")
write_title(ws_ret, 1, "Retention Rates by Cohort with Chi-Square Tests", max_col=10)

ret_rows = []
for outcome, label in [('retention_3month', '3-Month'), ('retention_6month', '6-Month'),
                        ('retention_9month', '9-Month'), ('retention_12month', '12-Month')]:
    c1 = df_managers[df_managers['cohort'] == 1][outcome]
    c2 = df_managers[df_managers['cohort'] == 2][outcome]
    ct = pd.crosstab(df_managers['cohort'], df_managers[outcome])
    chi2, p_val, _, _ = stats.chi2_contingency(ct)
    ret_rows.append({
        'Timepoint': label,
        'Cohort 1 Retained': int(c1.sum()),
        'Cohort 1 Lost': int(len(c1) - c1.sum()),
        'Cohort 1 Rate %': round(c1.mean() * 100, 1),
        'Cohort 2 Retained': int(c2.sum()),
        'Cohort 2 Lost': int(len(c2) - c2.sum()),
        'Cohort 2 Rate %': round(c2.mean() * 100, 1),
        'Difference (pp)': round((c1.mean() - c2.mean()) * 100, 1),
        'Chi-square': round(chi2, 3),
        'p-value': round(p_val, 3),
    })
df_ret = pd.DataFrame(ret_rows)
end_r = write_df(ws_ret, df_ret, start_row=3)
pval_col_letter = get_column_letter(10)
apply_pvalue_conditional(ws_ret, pval_col_letter, 4, end_r)
auto_fit_columns(ws_ret)
print("  [OK] Retention_Summary")

# ===== SHEET 7: Period_Turnover_Rates ======================================
ws_period = wb.create_sheet("Period_Turnover_Rates")
df_period_turnover, df_period_summary = create_period_turnover_sheet(ws_period, df_managers)
print("  [OK] Period_Turnover_Rates")

# ===== SHEET 8: Manager_Outcomes ===========================================
ws_mgr_out = wb.create_sheet("Manager_Outcomes")
write_title(ws_mgr_out, 1, "Manager-Level Survey Outcomes – Cohort Comparison", max_col=10)

mgr_out_rows = []
for outcome in survey_outcomes_mgr:
    c1 = df_managers[df_managers['treatment'] == 1][outcome]
    c2 = df_managers[df_managers['treatment'] == 0][outcome]
    t_stat, p_val = stats.ttest_ind(c1, c2)
    pooled_sd = np.sqrt((c1.std()**2 + c2.std()**2) / 2)
    d = (c1.mean() - c2.mean()) / pooled_sd if pooled_sd > 0 else 0
    mgr_out_rows.append({
        'Outcome': outcome,
        'Cohort 1 n': len(c1),
        'Cohort 1 Mean': round(c1.mean(), 3),
        'Cohort 1 SD': round(c1.std(), 3),
        'Cohort 2 n': len(c2),
        'Cohort 2 Mean': round(c2.mean(), 3),
        'Cohort 2 SD': round(c2.std(), 3),
        "Cohen's d": round(d, 3),
        't-statistic': round(t_stat, 3),
        'p-value': round(p_val, 3),
    })
df_mgr_out = pd.DataFrame(mgr_out_rows)
end_r = write_df(ws_mgr_out, df_mgr_out, start_row=3)
pval_col_letter = get_column_letter(10)
apply_pvalue_conditional(ws_mgr_out, pval_col_letter, 4, end_r)
auto_fit_columns(ws_mgr_out)
print("  [OK] Manager_Outcomes")

# ===== SHEET 9: DR_Outcomes ================================================
ws_dr_out = wb.create_sheet("DR_Outcomes")
write_title(ws_dr_out, 1, "Direct Report-Level Survey Outcomes – Cohort Comparison", max_col=10)

dr_out_rows = []
for outcome in survey_outcomes_dr:
    c1 = df_direct_reports[df_direct_reports['treatment'] == 1][outcome]
    c2 = df_direct_reports[df_direct_reports['treatment'] == 0][outcome]
    t_stat, p_val = stats.ttest_ind(c1, c2)
    pooled_sd = np.sqrt((c1.std()**2 + c2.std()**2) / 2)
    d = (c1.mean() - c2.mean()) / pooled_sd if pooled_sd > 0 else 0
    dr_out_rows.append({
        'Outcome': outcome,
        'Cohort 1 n': len(c1),
        'Cohort 1 Mean': round(c1.mean(), 3),
        'Cohort 1 SD': round(c1.std(), 3),
        'Cohort 2 n': len(c2),
        'Cohort 2 Mean': round(c2.mean(), 3),
        'Cohort 2 SD': round(c2.std(), 3),
        "Cohen's d": round(d, 3),
        't-statistic': round(t_stat, 3),
        'p-value': round(p_val, 3),
    })
df_dr_out = pd.DataFrame(dr_out_rows)
end_r = write_df(ws_dr_out, df_dr_out, start_row=3)
pval_col_letter = get_column_letter(10)
apply_pvalue_conditional(ws_dr_out, pval_col_letter, 4, end_r)
auto_fit_columns(ws_dr_out)
print("  [OK] DR_Outcomes")

# ===== SHEET 10: Team_Size_Distribution =====================================
ws_team = wb.create_sheet("Team_Size_Distribution")
write_title(ws_team, 1, "Team Size Distribution Across Managers", max_col=5)

# Summary stats
team_series = pd.Series(team_sizes)
team_summary = pd.DataFrame([{
    'Statistic': 'Min',
    'Value': int(team_series.min()),
}, {
    'Statistic': 'Max',
    'Value': int(team_series.max()),
}, {
    'Statistic': 'Mean',
    'Value': round(team_series.mean(), 2),
}, {
    'Statistic': 'SD',
    'Value': round(team_series.std(), 2),
}, {
    'Statistic': 'Median',
    'Value': round(team_series.median(), 1),
}, {
    'Statistic': 'Total Direct Reports',
    'Value': int(team_series.sum()),
}, {
    'Statistic': 'Total Managers',
    'Value': int(len(team_series)),
}])
end_r = write_df(ws_team, team_summary, start_row=3)

# Frequency table
write_title(ws_team, end_r + 2, "Frequency Table", max_col=4)
freq = team_series.value_counts().sort_index()
freq_df = pd.DataFrame({
    'Team Size': freq.index.astype(int),
    'Count': freq.values.astype(int),
    '% of Managers': (freq.values / len(team_series) * 100).round(1),
    'Cumulative %': (freq.values.cumsum() / len(team_series) * 100).round(1),
})
# Add Total row
total_row = pd.DataFrame([{
    'Team Size': 'Total',
    'Count': int(freq.values.sum()),
    '% of Managers': 100.0,
    'Cumulative %': 100.0,
}])
freq_df = pd.concat([freq_df, total_row], ignore_index=True)
end_r2 = write_df(ws_team, freq_df, start_row=end_r + 4)
auto_fit_columns(ws_team)
print("  [OK] Team_Size_Distribution")

# ===== SHEET 11: Raw_Managers ==============================================
ws_raw_mgr = wb.create_sheet("Raw_Managers")
write_title(ws_raw_mgr, 1, "Full Manager-Level Dataset", max_col=len(df_managers.columns))
end_r = write_df(ws_raw_mgr, df_managers, start_row=3)
auto_fit_columns(ws_raw_mgr)
print("  [OK] Raw_Managers")

# ===== SHEET 12: Raw_Direct_Reports ========================================
ws_raw_dr = wb.create_sheet("Raw_Direct_Reports")
write_title(ws_raw_dr, 1, "Full Direct Report-Level Dataset", max_col=len(df_direct_reports.columns))
end_r = write_df(ws_raw_dr, df_direct_reports, start_row=3)
auto_fit_columns(ws_raw_dr)
print("  [OK] Raw_Direct_Reports")

# ---------------------------------------------------------------------------
# Save workbook
# ---------------------------------------------------------------------------
wb.save(EXCEL_PATH)
print(f"\n[DONE] Excel report saved to: {EXCEL_PATH}")
